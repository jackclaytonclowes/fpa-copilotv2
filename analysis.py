"""
FP&A Copilot — pure analysis module.
All data-processing logic extracted from app.py; no Streamlit dependency.
"""
import json
import os
import re
import sys
import zipfile
from io import BytesIO
from pathlib import Path

import pandas as pd

# Set WATERFALL_DEBUG=1 to enable verbose per-period waterfall logging to stderr.
_WATERFALL_DEBUG = os.environ.get("WATERFALL_DEBUG", "").strip() == "1"

# ─────────────────────────────────────────────
# CLASSIFICATION CONFIG
# ─────────────────────────────────────────────
CATEGORY_RULES = {
    "Revenue": [
        "sales", "income", "revenue", "turnover", "fee income",
        "service income", "consultancy income", "grant income", "contract income",
        "funding", "uplift payment", "private income", "teaching income", "training hub",
        "qof", "global sum", "lcs", "des", "apms", "vaccination", "vaccinations",
        "influenza", "flu", "mmr", "rotavirus", "hpv", "rsv", "pertussis", "shingles",
        "teledermatology", "healthchecks", "health checks", "weight management",
        "access lis", "pcn", "arrs", "participation payment", "meeting backfill",
        "capacity and access fund", "supervision", "educational supervision", "grant",
        "gnrh", "academic funding", "winter planning", "advice & guidance",
    ],
    "Direct Costs": [
        "cost of sales", "direct cost", "direct costs", "subcontractor",
        "subcontractors", "delivery costs", "project costs", "materials", "consumables",
        "drugs and vaccines", "medical supplies", "vaccines",
    ],
    "Staff Costs": [
        "wage", "wages", "salary", "salaries", "payroll", "ni", "national insurance",
        "pension", "staff costs", "staff cost", "bonus", "bonuses", "overtime",
        "holiday pay", "training wages", "agency staff", "temporary staff",
        "staff welfare", "work force", "workforce", "receptionist", "receptionists",
        "practice nurse", "nurse lead", "nursing associate", "clinical lead",
        "pharmacist", "pharmacists", "salaried gp", "care navigator",
        "administrator", "senior administrator", "operational support",
        "operational lead", "practice managers", "trainee gp nurse", "hca",
        "doctors - clinical", "nurses clinical", "operations",
    ],
    "Management Costs": [
        "management", "management fee", "management fees", "manager", "leadership",
        "director", "directors", "board costs",
    ],
    "IT Costs": [
        "it", "it costs", "it + office equipment", "software", "hardware", "licence",
        "licences", "license", "licenses", "system", "systems", "tech", "technology",
        "computer", "computers", "microsoft", "office 365", "google workspace",
        "adobe", "xero", "sage", "subscription", "subscriptions", "server", "hosting",
        "domain", "website hosting", "internet", "broadband", "wifi", "telecoms",
        "telephone system", "telephone & internet", "office equipment",
    ],
    "Premises Costs": [
        "rent", "rates", "business rates", "service charge", "service charges",
        "utilities", "electricity", "gas", "water", "cleaning", "repairs",
        "maintenance", "security", "premises", "office rent", "building",
        "sewage", "drainage",
        "facilities", "room bookings", "room booking", "premises costs",
        "non-reimburseable premises costs", "non-reimbursable premises costs",
        "repair, renewals & maintenance",
    ],
    "Professional Fees": [
        "legal", "legal fees", "accountancy", "accounting", "audit", "auditor",
        "consultancy", "consultant", "professional fees", "advisor", "adviser",
        "legal expenses & professional fees",
    ],
    "Marketing Costs": [
        "marketing", "advertising", "promotion", "promotional", "branding",
        "seo", "google ads", "facebook ads", "meta ads", "campaign",
    ],
    "Travel & Entertainment": [
        "travel", "subsistence", "hotel", "hotels", "mileage", "train", "parking",
        "taxi", "taxis", "entertainment", "client entertainment",
    ],
    "Office & Admin": [
        "postage", "courier", "printing", "stationery", "office supplies",
        "admin", "administration", "general expenses", "sundry", "misc",
        "miscellaneous", "cqc costs", "postage, freight & courier",
    ],
    "Insurance": [
        "insurance", "liability insurance", "professional indemnity", "employers liability",
    ],
    "Finance Costs": [
        "bank charges", "interest", "loan interest", "finance charges", "merchant fees",
    ],
    "Depreciation & Amortisation": [
        "depreciation", "amortisation", "amortization",
    ],
    "Tax": ["corporation tax", "taxation", "tax"],
}

ACCOUNT_CATEGORY_OVERRIDES = {
    "global sum": "Revenue", "qof": "Revenue", "enhanced services": "Revenue",
    "private income": "Revenue", "training grant": "Revenue",
    "teaching income": "Revenue", "pcn des participation payment": "Revenue",
    "access lis funding": "Revenue", "winter planning": "Revenue",
    "teledermatology": "Revenue", "mmr": "Revenue", "rotavirus": "Revenue",
    "shingles": "Revenue", "flu": "Revenue", "vaccinations": "Revenue",
    "staff cost - practice care navigator": "Staff Costs",
    "staff cost - receptionist": "Staff Costs",
    "staff cost - salaried gp": "Staff Costs",
    "staff cost - pharmacist": "Staff Costs",
    "it + office equipment": "IT Costs",
    "telephone & internet": "IT Costs",
    "legal expenses & professional fees": "Professional Fees",
    "non-reimburseable premises costs": "Premises Costs",
    "non-reimbursable premises costs": "Premises Costs",
    "cqc costs": "Office & Admin",
    "cost of sales": "Direct Costs",
    "cost of goods sold": "Direct Costs",
    "cogs": "Direct Costs",
}

EXPENSE_CATEGORIES = [
    "Direct Costs", "Staff Costs", "Management Costs", "IT Costs",
    "Premises Costs", "Professional Fees", "Marketing Costs",
    "Travel & Entertainment", "Office & Admin", "Insurance",
    "Finance Costs", "Depreciation & Amortisation", "Tax", "Other",
]

SECTION_TO_CATEGORY = {
    "turnover": "Revenue", "revenue": "Revenue", "income": "Revenue",
    "administrative costs": "Other", "other costs": "Other",
    "staff & workforce": "Staff Costs", "staff and workforce": "Staff Costs",
}

CHART_COLORS_CSS = [
    "var(--c-1)", "var(--c-2)", "var(--c-3)", "var(--c-4)",
    "var(--c-5)", "var(--c-6)", "var(--c-7)", "var(--c-8)",
]


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fmt_gbp(v, sym="£"):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = sym.strip()
    return f"-{s}{abs(v):,.0f}" if v < 0 else f"{s}{v:,.0f}"


def fmt_pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return f"{v:.1f}%"


def clean_text(val):
    if pd.isna(val):
        return ""
    return str(val).strip()


def normalise(val):
    return clean_text(val).lower().strip()


def first_non_blank(values):
    for v in values:
        t = clean_text(v)
        if t and t.lower() != "nan":
            return t
    return ""


def is_heading_only_row(left, account):
    headings = {"turnover", "administrative costs", "other costs", "staff & workforce", "staff and workforce"}
    return normalise(left) in headings and normalise(account) == ""


def is_total_line(left, account):
    kws = [
        "total ", "gross profit", "operating profit", "operating surplus",
        "profit on ordinary activities", "profit before tax", "profit after tax",
        "net profit", "profit for the period", "surplus for the period",
        "net surplus", "ebitda", "ebit",
    ]
    return any(k in normalise(left) for k in kws) or any(k in normalise(account) for k in kws)


# ─────────────────────────────────────────────
# FILE LOADING
# ─────────────────────────────────────────────
def load_file(uploaded_bytes: bytes, filename: str) -> pd.DataFrame:
    if filename.endswith(".csv"):
        raw = pd.read_csv(BytesIO(uploaded_bytes), header=None)
    else:
        raw = pd.read_excel(BytesIO(uploaded_bytes), header=None)

    header_row = None
    for i in range(min(25, len(raw))):
        if "account" in raw.iloc[i].astype(str).str.strip().str.lower().tolist():
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find an 'Account' header row in the file.")

    headers = raw.iloc[header_row].astype(str).str.strip().tolist()
    df = raw.iloc[header_row + 1:].copy()
    df.columns = headers

    first_col = df.columns[0]
    account_col = next((c for c in df.columns if str(c).strip().lower() == "account"), None)
    if account_col is None:
        raise ValueError("Could not find the Account column.")

    df = df.rename(columns={first_col: "Level_1", account_col: "Level_2_Account"})
    month_cols = [c for c in df.columns if c not in ["Level_1", "Level_2_Account"]]

    section = ""
    rows = []
    for _, row in df.iterrows():
        left = clean_text(row.get("Level_1", ""))
        acc  = clean_text(row.get("Level_2_Account", ""))

        if is_heading_only_row(left, acc):
            section = left
            continue

        # For total/subtotal rows: prefer acc (e.g. "Total Turnover") over left (e.g. "Turnover"),
        # so that is_subtotal() can correctly identify them by their full name.
        # For normal rows: prefer acc (the specific account name) over left (the section heading).
        # Both branches are identical — always prefer the more specific acc column first.
        account_name = first_non_blank([acc, left])
        if not account_name:
            continue

        r = {"Account": account_name, "Section": section}
        for col in month_cols:
            r[col] = row[col]
        rows.append(r)

    return pd.DataFrame(rows)


def clean_numeric(series):
    return pd.to_numeric(
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("£", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
        .str.replace(r"[^\d.\-]", "", regex=True)
        .str.strip()
        .replace("", pd.NA),
        errors="coerce",
    )


# ─────────────────────────────────────────────
# CLASSIFICATION
# ─────────────────────────────────────────────
def _kw_matches(kw: str, acc: str) -> bool:
    if any(c in kw for c in (" ", "-", "&")):
        return kw in acc
    return bool(re.search(r'\b' + re.escape(kw) + r'\b', acc))


_CLASSIFICATIONS_DIR = Path(__file__).parent / "classifications"


def load_sector_synonyms(sector: str) -> dict[str, str]:
    """Load account→category synonym overrides for a given sector.

    Returns an empty dict for unknown/unconfigured sectors so callers
    don't need to guard against None.
    """
    path = _CLASSIFICATIONS_DIR / f"{sector}.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return {k.lower().strip(): v for k, v in data.get("synonyms", {}).items()}
    except Exception:
        return {}


def classify(account, section=None, extra_overrides: dict | None = None):
    acc = normalise(account)
    sec = normalise(section) if section else ""

    # Sector-specific synonyms checked before built-in rules
    if extra_overrides:
        if acc in extra_overrides:
            return extra_overrides[acc]
        for k, v in extra_overrides.items():
            if k in acc:
                return v

    if acc in ACCOUNT_CATEGORY_OVERRIDES:
        return ACCOUNT_CATEGORY_OVERRIDES[acc]
    for k, v in ACCOUNT_CATEGORY_OVERRIDES.items():
        if k in acc:
            return v
    if sec in SECTION_TO_CATEGORY:
        sc = SECTION_TO_CATEGORY[sec]
        if sc in ("Revenue", "Staff Costs"):
            return sc
    for cat, kws in CATEGORY_RULES.items():
        if any(_kw_matches(kw, acc) for kw in kws):
            return cat
    return "Other"


def is_subtotal(account):
    kws = [
        "total ", "gross profit", "operating profit", "operating surplus",
        "profit on ordinary activities", "profit before tax", "profit after tax",
        "net profit", "profit for the period", "surplus for the period",
        "net surplus", "ebitda", "ebit",
    ]
    return any(k in normalise(account) for k in kws)


# ─────────────────────────────────────────────
# PERIOD LOGIC
# ─────────────────────────────────────────────
def get_fy(dt):
    dt = pd.Timestamp(dt)
    return dt.year if dt.month >= 4 else dt.year - 1


def get_quarter(dt):
    dt = pd.Timestamp(dt)
    q = {4:"Q1",5:"Q1",6:"Q1",7:"Q2",8:"Q2",9:"Q2",10:"Q3",11:"Q3",12:"Q3",1:"Q4",2:"Q4",3:"Q4"}[dt.month]
    fy = get_fy(dt)
    return f"FY{str(fy)[-2:]}/{str(fy+1)[-2:]} {q}"


def quarter_sort_key(label):
    parts = label.split(" ")
    fy_start = int("20" + parts[0][2:4])
    qo = {"Q1":1,"Q2":2,"Q3":3,"Q4":4}[parts[1]]
    return fy_start * 10 + qo


def period_label(period_val, mode):
    if pd.isna(period_val):
        return ""
    if mode == "monthly":
        return pd.Timestamp(period_val).strftime("%B %Y")
    return str(period_val)


# ─────────────────────────────────────────────
# CORE BUILD
# ─────────────────────────────────────────────
def build_long(df: pd.DataFrame, mode: str = "monthly") -> pd.DataFrame:
    possible = [c for c in df.columns if c not in ["Account", "Section"]]
    month_cols = []
    for col in possible:
        parsed = pd.to_datetime(str(col).replace("Sept", "Sep").strip(), errors="coerce")
        if pd.notna(parsed):
            month_cols.append(col)

    if not month_cols:
        raise ValueError("No valid month columns detected.")

    df_long = df.melt(id_vars=["Account","Section"], value_vars=month_cols, var_name="Month", value_name="Value")
    df_long["Month"] = pd.to_datetime(
        df_long["Month"].astype(str).str.replace("Sept", "Sep", regex=False).str.strip(), errors="coerce"
    )
    df_long["Value"] = clean_numeric(df_long["Value"])
    df_long = df_long.dropna(subset=["Month","Value"])
    df_long = df_long.groupby(["Account","Section","Month"], as_index=False)["Value"].sum()

    if mode == "monthly":
        df_long["Period"] = df_long["Month"]
        df_long["Period Sort"] = df_long["Month"]
    else:
        df_long["Period"] = df_long["Month"].apply(get_quarter)
        df_long["Period Sort"] = df_long["Period"].apply(quarter_sort_key)
        df_long = df_long.groupby(["Account","Section","Period","Period Sort"], as_index=False)["Value"].sum()

    return df_long


def build_analysis(df_long: pd.DataFrame, extra_overrides: dict | None = None) -> pd.DataFrame:
    grp = df_long.copy()
    grp = grp.sort_values(["Account","Period Sort"])
    grp["Prior Value"] = grp.groupby("Account")["Value"].shift(1)
    grp["Variance"] = grp["Value"] - grp["Prior Value"]
    grp["Variance %"] = grp.apply(
        lambda r: (r["Variance"] / r["Prior Value"] * 100)
        if pd.notna(r["Prior Value"]) and r["Prior Value"] != 0 else None, axis=1
    )
    grp["Abs Variance"] = grp["Variance"].abs()
    grp["Category"] = grp.apply(
        lambda r: classify(r["Account"], r["Section"], extra_overrides), axis=1
    )
    grp["Is Subtotal"] = grp["Account"].apply(is_subtotal)
    return grp


# ─────────────────────────────────────────────
# KPI DETECTION
# ─────────────────────────────────────────────
def detect_kpis(df_long: pd.DataFrame) -> dict:
    all_accounts = df_long["Account"].dropna().astype(str).str.strip().unique().tolist()
    subtotals = [a for a in all_accounts if is_subtotal(a)]

    revenue_kws = ["total turnover","total revenue","total income","gross income","total sales","total fees","total funding"]
    cost_kws    = ["total administrative costs","total admin costs","total overheads","total expenses","total costs","total expenditure"]
    profit_kws  = ["operating profit","operating surplus","net profit","gross profit","profit before tax",
                   "profit for the period","ebitda","ebit","surplus for the period","net surplus"]

    def best(candidates, kws):
        for kw in kws:
            for a in candidates:
                if a.lower().strip() == kw:
                    return a
        for kw in kws:
            for a in candidates:
                if kw in a.lower():
                    return a
        return None

    return {
        "revenue": best(subtotals, revenue_kws) or best(all_accounts, revenue_kws),
        "costs":   best(subtotals, cost_kws)    or best(all_accounts, cost_kws),
        "profit":  best(subtotals, profit_kws)  or best(all_accounts, profit_kws),
    }


# ─────────────────────────────────────────────
# WATERFALL ANALYSIS
# ─────────────────────────────────────────────
def build_waterfall(
    analysis_df: pd.DataFrame,
    selected_period,
    kpi_accounts: dict,
    mode: str,
) -> dict | None:
    """
    Compute what drove the change in Operating Profit period-on-period.

    Returns a dict with:
        prior_profit, current_profit, net_change,
        bars: [{label, impact, fav}],  (sorted Revenue-first, then by |impact|)
        largest_positive / largest_negative,
        commentary: str

    Returns None when no prior period exists (first period in file).
    """
    profit_account = kpi_accounts.get("profit")
    if not profit_account:
        return None

    period_df = analysis_df[analysis_df["Period"] == selected_period].copy()

    # ── Prior / current profit from the KPI subtotal row ─────────────────
    prof_row = period_df[period_df["Account"].str.strip() == profit_account.strip()]
    if prof_row.empty:
        return None

    r = prof_row.iloc[0]
    current_profit = float(r["Value"])      if pd.notna(r["Value"])      else 0.0
    prior_profit   = float(r["Prior Value"]) if pd.notna(r["Prior Value"]) else None
    if prior_profit is None:
        return None

    net_change = current_profit - prior_profit

    # ── Per-account profit impact ─────────────────────────────────────────
    #
    # Exact required formula:
    #
    #   Revenue / income accounts:
    #       profit_impact = current_value − prior_value
    #       (revenue ↑  →  positive impact on profit)
    #
    #   Expense / cost accounts:
    #       profit_impact = prior_value − current_value
    #       (cost ↑  →  negative impact on profit)
    #       (cost ↓  →  positive impact on profit)
    #
    # Reconciliation identity (standard P&L, costs positive):
    #   Σ(rev: cur−pri) + Σ(cost: pri−cur)
    #   = (cur_rev_total − pri_rev_total) − (cur_cost_total − pri_cost_total)
    #   = (cur_rev − cur_cost) − (pri_rev − pri_cost)
    #   = current_profit − prior_profit  ✓
    #
    # Using Value/Prior Value directly (not the pre-computed Variance column)
    # avoids any DataFrame computation artifact.
    #
    # Revenue detection: rely solely on Category == "Revenue".
    # classify() already applies SECTION_TO_CATEGORY when the account's
    # Section is a known income section, so we trust that result here.
    # Adding a second Section check inside this function was the previous bug —
    # it could treat cost accounts that happen to share a section name as
    # revenue and flip their sign.

    def _is_revenue(row) -> bool:
        return row.get("Category") == "Revenue"

    def _profit_impact(row) -> float:
        cur = float(row["Value"])       if pd.notna(row.get("Value"))       else 0.0
        pri = float(row["Prior Value"]) if pd.notna(row.get("Prior Value")) else 0.0
        if _is_revenue(row):
            return cur - pri   # Revenue ↑  →  profit ↑  (positive)
        else:
            return pri - cur   # Cost    ↑  →  profit ↓  (negative)

    driver_df = period_df[
        (~period_df["Is Subtotal"]) &
        period_df["Prior Value"].notna()   # need both values to compute impact
    ].copy()

    driver_df["ProfitImpact"] = driver_df.apply(_profit_impact, axis=1)

    # ── Aggregate by category — capture source values for diagnostics ────────
    cat_agg = (
        driver_df.groupby("Category")
        .agg(
            impact       =("ProfitImpact", "sum"),
            current_total=("Value",        lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum()),
            prior_total  =("Prior Value",  lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum()),
        )
        .reset_index()
    )

    # ── Diagnostic: print source values to server stderr (opt-in via WATERFALL_DEBUG=1) ──
    if _WATERFALL_DEBUG:
        print(f"\n[waterfall] ── {selected_period} ──────────────────────────────────", file=sys.stderr)
        print(f"[waterfall]   Prior profit : {fmt_gbp(prior_profit)}", file=sys.stderr)
        print(f"[waterfall]   Current profit: {fmt_gbp(current_profit)}", file=sys.stderr)
        print(f"[waterfall]   Net change   : {'+' if net_change >= 0 else ''}{fmt_gbp(net_change)}", file=sys.stderr)
        print(f"[waterfall]   {'Category':<30} {'Prior':>12} {'Current':>12} {'RawVariance':>13} {'ProfitImpact':>13} Type", file=sys.stderr)
        for _, row in cat_agg.iterrows():
            cat  = str(row["Category"])
            cur  = float(row["current_total"])
            pri  = float(row["prior_total"])
            pi   = float(row["impact"])
            raw  = cur - pri
            kind = "revenue (cur−pri)" if _is_revenue({"Category": cat}) else "cost (pri−cur)"
            print(
                f"[waterfall]   {cat:<30} {fmt_gbp(pri):>12} {fmt_gbp(cur):>12} "
                f"{('+' if raw>=0 else '')}{fmt_gbp(raw):>13} "
                f"{('+' if pi>=0 else '')}{fmt_gbp(pi):>13} {kind}",
                file=sys.stderr,
            )

    # ── category_detail for API response / frontend console ──────────────────
    category_detail = [
        {
            "category"      : str(row["Category"]),
            "prior"         : round(float(row["prior_total"]), 2),
            "current"       : round(float(row["current_total"]), 2),
            "raw_variance"  : round(float(row["current_total"]) - float(row["prior_total"]), 2),
            "profit_impact" : round(float(row["impact"]), 2),
            "is_revenue"    : _is_revenue({"Category": str(row["Category"])}),
        }
        for _, row in cat_agg.iterrows()
    ]

    cat_df = cat_agg[["Category", "impact"]].copy()
    cat_df = cat_df[cat_df["impact"].abs().round(2) > 0]

    # ── Materiality filter: keep categories > 5 % of |net_change| ─────────
    threshold = abs(net_change) * 0.05 if abs(net_change) > 0 else 0

    if threshold > 0:
        material   = cat_df[cat_df["impact"].abs() >= threshold].copy()
        immaterial = cat_df[cat_df["impact"].abs() <  threshold].copy()
    else:
        material   = cat_df.copy()
        immaterial = pd.DataFrame(columns=cat_df.columns)

    if not immaterial.empty:
        other_sum = float(immaterial["impact"].sum())
        if abs(other_sum) > 0.01:
            material = pd.concat(
                [material, pd.DataFrame([{"Category": "Other Movements", "impact": other_sum}])],
                ignore_index=True,
            )

    # ── Sort: Revenue first, then by |impact| desc, Other last ────────────
    def _sort_key(row):
        cat = row["Category"]
        if cat == "Revenue":            return (0, -abs(row["impact"]))
        if cat == "Other Movements":    return (99, 0)
        return                                 (1,  -abs(row["impact"]))

    material["_sk"] = material.apply(_sort_key, axis=1)
    material = material.sort_values("_sk").drop(columns=["_sk"])

    bars = [
        {"label": r["Category"], "impact": round(float(r["impact"]), 2), "fav": float(r["impact"]) > 0}
        for _, r in material.iterrows()
    ]

    # ── Reconciliation guarantee ──────────────────────────────────────────
    # The sum of all bar impacts MUST equal net_change.
    # Any residual (from mis-classification, rounding, or accounts missing
    # from the category buckets) is surfaced explicitly rather than hidden.
    # ── Reconciliation check ──────────────────────────────────────────────
    # prior_profit + sum(bar impacts) must equal current_profit exactly.
    # Any residual is surfaced as "Other Movements"; a server-side warning
    # is printed and reconciles=False is returned to the frontend.
    explained  = round(sum(b["impact"] for b in bars), 2)
    residual   = round(net_change - explained, 2)
    reconciles = abs(residual) <= 0.5

    if not reconciles:
        print(
            f"[waterfall] RECONCILIATION MISMATCH  "
            f"net_change={net_change:+.2f}  explained={explained:+.2f}  "
            f"residual={residual:+.2f}  period={selected_period}",
            file=sys.stderr,
        )
        other_idx = next((i for i, b in enumerate(bars) if b["label"] == "Other Movements"), None)
        if other_idx is not None:
            merged = round(bars[other_idx]["impact"] + residual, 2)
            bars[other_idx] = {"label": "Other Movements", "impact": merged, "fav": merged > 0}
        else:
            bars.append({"label": "Other Movements", "impact": residual, "fav": residual > 0})

    # ── Largest drivers ───────────────────────────────────────────────────
    non_other = [b for b in bars if b["label"] != "Other Movements"]
    pos_bars  = [b for b in non_other if b["impact"] > 0]
    neg_bars  = [b for b in non_other if b["impact"] < 0]
    largest_pos = max(pos_bars, key=lambda b: b["impact"]) if pos_bars else None
    largest_neg = min(neg_bars, key=lambda b: b["impact"]) if neg_bars else None

    # ── Commentary — profit-impact wording ────────────────────────────────
    # Every sentence describes effect on profit, not raw account movement.
    # "Staff Costs reduced profit by £4,479 due to higher costs"   (correct)
    # "Staff Costs improved profit by £7,231 due to lower costs"   (correct)
    direction = "increased" if net_change > 0 else "decreased"
    parts = [
        f"{profit_account} {direction} by {fmt_gbp(abs(net_change))} "
        f"compared with the prior period."
    ]

    def _cat_sentence(bar: dict) -> str:
        cat    = bar["label"]
        impact = bar["impact"]
        fav    = impact > 0
        if cat == "Revenue":
            move = "increased" if fav else "decreased"
            effect = "contributing positively to" if fav else "reducing"
            return f"Revenue {move} by {fmt_gbp(abs(impact))}, {effect} profit."
        if cat == "Other Movements":
            effect = "improved" if fav else "reduced"
            return f"Other movements {effect} profit by {fmt_gbp(abs(impact))}."
        # Named cost category
        if fav:
            return (f"{cat} improved profit by {fmt_gbp(abs(impact))} "
                    f"due to lower costs in the period.")
        return (f"{cat} reduced profit by {fmt_gbp(abs(impact))} "
                f"due to higher costs in the period.")

    if largest_pos and largest_neg:
        rev_led = largest_pos["label"] == "Revenue"
        if rev_led and net_change < 0:
            parts.append(
                f"Revenue increased by {fmt_gbp(largest_pos['impact'])} but was insufficient "
                f"to offset {largest_neg['label']}, which reduced profit by "
                f"{fmt_gbp(abs(largest_neg['impact']))} due to higher costs."
            )
        elif rev_led:
            parts.append(
                f"Revenue growth of {fmt_gbp(largest_pos['impact'])} was the primary driver. "
                f"{largest_neg['label']} partially offset this, reducing profit by "
                f"{fmt_gbp(abs(largest_neg['impact']))} due to higher costs."
            )
        else:
            parts.append(_cat_sentence(largest_pos))
            parts.append(_cat_sentence(largest_neg))
    elif largest_neg:
        parts.append(_cat_sentence(largest_neg))
    elif largest_pos:
        parts.append(_cat_sentence(largest_pos))

    return {
        "prior_profit":     round(prior_profit, 2),
        "current_profit":   round(current_profit, 2),
        "net_change":       round(net_change, 2),
        "bars":             bars,
        "largest_positive": largest_pos,
        "largest_negative": largest_neg,
        "commentary":       " ".join(parts),
        "profit_account":   profit_account,
        "reconciles":       reconciles,
        "category_detail":  category_detail,   # full source values for inspection
    }


# ─────────────────────────────────────────────
# PERIOD DATA FOR API RESPONSE
# ─────────────────────────────────────────────
def get_period_data(analysis_df: pd.DataFrame, df_long: pd.DataFrame,
                    selected_period, kpi_accounts: dict, mode: str) -> dict:
    period_df  = analysis_df[analysis_df["Period"] == selected_period].copy()
    driver_df  = period_df[~period_df["Is Subtotal"]].copy()

    # KPI rows
    def kpi_row(name, fav_when_up, icon):
        if not name:
            return None
        match = period_df[period_df["Account"].str.strip() == name.strip()]
        if match.empty:
            return None
        r = match.iloc[0]
        variance = float(r["Variance"]) if pd.notna(r["Variance"]) else None
        prior    = float(r["Prior Value"]) if pd.notna(r["Prior Value"]) else None
        value    = float(r["Value"]) if pd.notna(r["Value"]) else None
        pct      = float(r["Variance %"]) if pd.notna(r["Variance %"]) else None
        is_fav   = (fav_when_up and variance >= 0) or (not fav_when_up and variance <= 0) if variance is not None else None
        return {"label": name, "value": value, "variance": variance, "prior": prior, "pct": pct,
                "is_fav": is_fav, "icon": icon, "pct_only": False}

    rev_kpi  = kpi_row(kpi_accounts["revenue"], True,  "trending-up")
    cost_kpi = kpi_row(kpi_accounts["costs"],   False, "receipt")
    prof_kpi = kpi_row(kpi_accounts["profit"],  True,  "wallet")

    kpis = []
    for k in [rev_kpi, cost_kpi, prof_kpi]:
        if k:
            kpis.append(k)
    # Profit % card
    if prof_kpi and prof_kpi["pct"] is not None:
        kpis.append({"label": "Profit Variance", "value": None, "variance": prof_kpi["variance"],
                     "pct": prof_kpi["pct"], "is_fav": prof_kpi["is_fav"], "icon": "bar-chart-2", "pct_only": True})

    # Trend (last 6 months in long form)
    sorted_periods = sorted(analysis_df["Period"].unique(),
        key=lambda p: pd.Timestamp(p) if mode == "monthly" else quarter_sort_key(p))
    idx = list(sorted_periods).index(selected_period) if selected_period in sorted_periods else len(sorted_periods)-1
    trend_periods = sorted_periods[max(0, idx-5):idx+1]

    trend_data = []
    for p in trend_periods:
        lbl = period_label(p, mode)
        short = pd.Timestamp(p).strftime("%b") if mode == "monthly" else str(p)[-5:]
        def val(name):
            if not name:
                return 0
            m = analysis_df[(analysis_df["Period"] == p) & (analysis_df["Account"].str.strip() == name.strip())]
            return float(m.iloc[0]["Value"]) if not m.empty and pd.notna(m.iloc[0]["Value"]) else 0
        trend_data.append({
            "m": short, "full": lbl,
            "revenue": val(kpi_accounts["revenue"]),
            "costs":   val(kpi_accounts["costs"]),
            "profit":  val(kpi_accounts["profit"]),
        })

    # Revenue & expense splits
    def expense_split_fn(categories, top_n=6):
        """Group expenses by Category for a meaningful cost-category breakdown."""
        df = driver_df[driver_df["Category"].isin(categories)].copy()
        df = df.groupby("Category", as_index=False)["Value"].sum()
        df["Chart Value"] = df["Value"].abs()
        df = df[df["Chart Value"] > 0].sort_values("Chart Value", ascending=False)
        if len(df) > top_n:
            other = df.iloc[top_n:]["Chart Value"].sum()
            top   = df.head(top_n).copy()
            if other > 0:
                df = pd.concat([top, pd.DataFrame([{"Category": "Other", "Value": other, "Chart Value": other}])], ignore_index=True)
            else:
                df = top
        total = df["Chart Value"].sum()
        return [{"name": r["Category"], "value": float(r["Chart Value"]),
                 "pct": round(r["Chart Value"] / total * 100, 1) if total else 0}
                for _, r in df.iterrows()]

    def revenue_split_fn(top_n=8):
        """Group revenue by individual Account so we can see the composition of income streams.

        Rules:
        - Only rows classified as Revenue (Category == "Revenue")
        - driver_df already excludes Is Subtotal rows (Total Turnover, Gross Profit, etc.)
        - Exclude any account whose name looks like a subtotal/total line just in case
        - Group by Account name, show positive values only
        - If only one account survives, fall back to that single row (still useful, not a bug)
        """
        # Known section-heading names that appear as account names when a total row
        # uses the section name (e.g. "Turnover" as a stand-in for "Total Turnover").
        SECTION_HEADING_NAMES = {"turnover", "income", "revenue", "receipts"}
        SUBTOTAL_PATTERNS = [
            "total", "gross profit", "subtotal", "sub-total",
            "net revenue", "net income", "total income",
        ]

        def looks_like_subtotal(name: str) -> bool:
            n = name.lower().strip()
            if n in SECTION_HEADING_NAMES:
                return True
            return any(p in n for p in SUBTOTAL_PATTERNS)

        df = driver_df[
            (driver_df["Category"] == "Revenue") &
            (~driver_df["Account"].apply(looks_like_subtotal))
        ].copy()

        # Group by Account (individual income lines)
        df = df.groupby("Account", as_index=False)["Value"].sum()
        df["Chart Value"] = df["Value"].abs()
        df = df[df["Chart Value"] > 0].sort_values("Chart Value", ascending=False)

        if df.empty:
            return []

        if len(df) > top_n:
            other = df.iloc[top_n:]["Chart Value"].sum()
            top   = df.head(top_n).copy()
            if other > 0:
                df = pd.concat([top, pd.DataFrame([{"Account": "Other", "Value": other, "Chart Value": other}])], ignore_index=True)
            else:
                df = top

        total = df["Chart Value"].sum()
        return [{"name": r["Account"], "value": float(r["Chart Value"]),
                 "pct": round(r["Chart Value"] / total * 100, 1) if total else 0}
                for _, r in df.iterrows()]

    revenue_split = revenue_split_fn()
    expense_split = expense_split_fn(EXPENSE_CATEGORIES)

    # Per-account value history for sparklines (up to last 12 periods, in order)
    spark_history: dict[str, list] = {}
    for p in sorted_periods[-12:]:
        p_df = analysis_df[analysis_df["Period"] == p]
        for _, srow in p_df[~p_df["Is Subtotal"]].iterrows():
            acc = srow["Account"].strip()
            v = float(srow["Value"]) if pd.notna(srow["Value"]) else None
            spark_history.setdefault(acc, []).append(v)

    # Top movements
    top_mov = (driver_df[driver_df["Variance"].notna() & (driver_df["Variance"] != 0)]
               .sort_values("Abs Variance", ascending=False).head(15))
    movements = []
    for _, row in top_mov.iterrows():
        variance = float(row["Variance"]) if pd.notna(row["Variance"]) else None
        cat = row["Category"]
        fav_when_up = cat == "Revenue"
        is_fav = (fav_when_up and variance >= 0) or (not fav_when_up and variance <= 0) if variance is not None else None
        movements.append({
            "account":     row["Account"],
            "category":    cat,
            "value":       float(row["Value"]) if pd.notna(row["Value"]) else None,
            "prior_value": float(row["Prior Value"]) if pd.notna(row["Prior Value"]) else None,
            "variance":    variance,
            "variance_pct": float(row["Variance %"]) if pd.notna(row["Variance %"]) else None,
            "is_fav":      is_fav,
            "history":     spark_history.get(row["Account"].strip(), []),
        })

    # Commentary
    commentary = []
    if prof_kpi and prof_kpi["variance"] is not None:
        word = "improved" if prof_kpi["variance"] > 0 else "declined"
        icon = "trending-up" if prof_kpi["variance"] > 0 else "trending-down"
        fav  = prof_kpi["variance"] > 0
        commentary.append({"icon": icon, "fav": fav,
            "html": f"<b>{kpi_accounts['profit'] or 'Operating profit'} {word}</b> by "
                    f"<b>{fmt_gbp(abs(prof_kpi['variance']))}</b>"
                    f"{(' (' + fmt_pct(abs(prof_kpi['pct'])) + ')') if prof_kpi['pct'] else ''} vs prior period.",
            "source": {"account": kpi_accounts['profit'] or 'Operating profit',
                       "actual": prof_kpi["value"], "actual_label": "Current",
                       "comparison": prof_kpi["prior"], "comparison_label": "Prior period",
                       "variance": prof_kpi["variance"], "pct": prof_kpi["pct"]}})
    if rev_kpi and rev_kpi["variance"]:
        d = "increased" if rev_kpi["variance"] > 0 else "decreased"
        commentary.append({"icon": "arrow-up-right" if rev_kpi["variance"]>0 else "arrow-down-right",
            "fav": rev_kpi["variance"] > 0,
            "html": f"Revenue {d} by <b>{fmt_gbp(abs(rev_kpi['variance']))}"
                    f"{(' (' + fmt_pct(abs(rev_kpi['pct'])) + ')') if rev_kpi['pct'] else ''}</b> vs prior period.",
            "source": {"account": rev_kpi["label"],
                       "actual": rev_kpi["value"], "actual_label": "Current",
                       "comparison": rev_kpi["prior"], "comparison_label": "Prior period",
                       "variance": rev_kpi["variance"], "pct": rev_kpi["pct"]}})
    for _, row in (driver_df[driver_df["Variance"].notna()]
                   .sort_values("Abs Variance", ascending=False).head(3).iterrows()):
        if row["Variance"] == 0:
            continue
        d = "increased" if row["Variance"] > 0 else "decreased"
        cat = row["Category"]
        fav = (cat == "Revenue" and row["Variance"] > 0) or (cat != "Revenue" and row["Variance"] < 0)
        pct_str = f" ({fmt_pct(abs(row['Variance %']))})" if pd.notna(row["Variance %"]) else ""
        commentary.append({
            "icon": "arrow-up-right" if row["Variance"] > 0 else "arrow-down-right",
            "fav": fav,
            "html": f"<b>{row['Account']}</b> {d} by {fmt_gbp(abs(row['Variance']))}{pct_str}.",
            "source": {"account": row["Account"], "category": cat,
                       "actual": float(row["Value"]) if pd.notna(row["Value"]) else None, "actual_label": "Current",
                       "comparison": float(row["Prior Value"]) if pd.notna(row["Prior Value"]) else None, "comparison_label": "Prior period",
                       "variance": float(row["Variance"]), "pct": float(row["Variance %"]) if pd.notna(row["Variance %"]) else None},
        })

    # Period label + prior
    selected_label = period_label(selected_period, mode)
    prior_period   = None
    if idx > 0:
        prior_period = period_label(sorted_periods[idx - 1], mode)

    waterfall = build_waterfall(analysis_df, selected_period, kpi_accounts, mode)

    return {
        "kpis":          kpis,
        "trend":         trend_data,
        "revenue_split": revenue_split,
        "expense_split": expense_split,
        "movements":     movements,
        "commentary":    commentary,
        "waterfall":     waterfall,
        "period":        {"label": selected_label, "prior": prior_period or "prior period"},
        "selected_period": str(selected_period),
        "periods":       [str(p) for p in sorted_periods],
    }


def get_ytd_data(analysis_df: pd.DataFrame, selected_period, kpi_accounts: dict) -> dict:
    """
    Year-to-date aggregate: sum from Jan (or first available month of the year)
    to selected_period.  Prior = same month range of the immediately preceding
    calendar year in the dataset (empty DataFrame if not available).
    Returns the same data dict structure as get_period_data().
    """
    all_periods = sorted(analysis_df["Period"].unique(), key=lambda p: pd.Timestamp(p))
    sel_ts = next((p for p in all_periods if str(p)[:10] == str(selected_period)[:10]), all_periods[-1])
    sel_dt = pd.Timestamp(sel_ts)

    ytd_curr  = [p for p in all_periods
                 if pd.Timestamp(p).year == sel_dt.year and pd.Timestamp(p) <= sel_ts]
    prior_year = sel_dt.year - 1
    ytd_prior  = [p for p in all_periods
                  if pd.Timestamp(p).year == prior_year and pd.Timestamp(p).month <= sel_dt.month]

    has_prior = len(ytd_prior) > 0

    def _agg(periods):
        if not periods:
            return pd.DataFrame(columns=["Account","Category","Is Subtotal","Value"])
        return (analysis_df[analysis_df["Period"].isin(periods)]
                .groupby(["Account","Category","Is Subtotal"], as_index=False)["Value"].sum())

    curr_agg  = _agg(ytd_curr)
    prior_agg = _agg(ytd_prior)

    merged = curr_agg.merge(
        prior_agg.rename(columns={"Value": "Prior Value"})[["Account","Category","Prior Value"]],
        on=["Account","Category"], how="left",
    )
    merged["Prior Value"]  = merged.get("Prior Value", pd.Series(0, index=merged.index)).fillna(0)
    merged["Variance"]     = merged["Value"] - merged["Prior Value"]
    merged["Variance %"]   = merged.apply(
        lambda r: r["Variance"] / r["Prior Value"] * 100 if r["Prior Value"] != 0 else None, axis=1)
    merged["Abs Variance"] = merged["Variance"].abs()

    driver_df = merged[~merged["Is Subtotal"]].copy()

    # ── KPIs ──────────────────────────────────────────────────────────────────
    def kpi_row(name, fav_when_up, icon):
        if not name:
            return None
        m = merged[merged["Account"].str.strip() == name.strip()]
        if m.empty:
            return None
        r = m.iloc[0]
        variance = float(r["Variance"]) if pd.notna(r["Variance"]) else None
        prior    = float(r["Prior Value"]) if pd.notna(r["Prior Value"]) else None
        value    = float(r["Value"]) if pd.notna(r["Value"]) else None
        pct      = float(r["Variance %"]) if pd.notna(r["Variance %"]) else None
        is_fav   = (fav_when_up and variance >= 0) or (not fav_when_up and variance <= 0) if variance is not None else None
        return {"label": name, "value": value, "variance": variance if has_prior else None,
                "prior": prior if has_prior else None, "pct": pct if has_prior else None,
                "is_fav": is_fav, "icon": icon, "pct_only": False}

    rev_kpi  = kpi_row(kpi_accounts.get("revenue"), True,  "trending-up")
    cost_kpi = kpi_row(kpi_accounts.get("costs"),   False, "receipt")
    prof_kpi = kpi_row(kpi_accounts.get("profit"),  True,  "wallet")
    kpis     = [k for k in [rev_kpi, cost_kpi, prof_kpi] if k]
    if prof_kpi and prof_kpi.get("pct") is not None:
        kpis.append({"label": "Profit YoY", "value": None,
                     "variance": prof_kpi["variance"], "pct": prof_kpi["pct"],
                     "is_fav": prof_kpi["is_fav"], "icon": "bar-chart-2", "pct_only": True})

    # ── Movements ──────────────────────────────────────────────────────────────
    top_mov = (driver_df[driver_df["Abs Variance"] > 0]
               .sort_values("Abs Variance", ascending=False).head(15))
    movements = []
    for _, row in top_mov.iterrows():
        variance = float(row["Variance"]) if pd.notna(row["Variance"]) else None
        cat      = row["Category"]
        fav_when_up = cat == "Revenue"
        is_fav   = (fav_when_up and variance >= 0) or (not fav_when_up and variance <= 0) if variance is not None else None
        movements.append({
            "account":      row["Account"],
            "category":     cat,
            "value":        float(row["Value"]) if pd.notna(row["Value"]) else None,
            "prior_value":  float(row["Prior Value"]) if pd.notna(row["Prior Value"]) else None,
            "variance":     variance if has_prior else None,
            "variance_pct": float(row["Variance %"]) if pd.notna(row["Variance %"]) else None,
            "is_fav":       is_fav,
            "history":      [],
        })

    # ── Commentary ─────────────────────────────────────────────────────────────
    commentary = []
    if prof_kpi and prof_kpi.get("variance") is not None:
        word = "ahead of" if prof_kpi["variance"] > 0 else "behind"
        fav  = prof_kpi["variance"] > 0
        commentary.append({"icon": "trending-up" if fav else "trending-down", "fav": fav,
            "html": f"<b>{kpi_accounts.get('profit','Operating Profit')} YTD</b> is "
                    f"{fmt_gbp(abs(prof_kpi['variance']))} {word} prior year YTD."})
    for _, row in (driver_df[driver_df["Abs Variance"] > 0]
                   .sort_values("Abs Variance", ascending=False).head(3).iterrows()):
        d   = "increased" if row["Variance"] > 0 else "decreased"
        cat = row["Category"]
        fav = (cat == "Revenue" and row["Variance"] > 0) or (cat != "Revenue" and row["Variance"] < 0)
        pct_str = f" ({fmt_pct(abs(row['Variance %']))})" if pd.notna(row["Variance %"]) else ""
        commentary.append({
            "icon": "arrow-up-right" if row["Variance"] > 0 else "arrow-down-right", "fav": fav,
            "html": f"<b>{row['Account']}</b> YTD {d} by {fmt_gbp(abs(row['Variance']))}{pct_str} vs prior year.",
        })

    months_n  = len(ytd_curr)
    end_month = sel_dt.strftime("%b %Y")
    ytd_label = f"YTD {end_month} ({months_n} month{'s' if months_n != 1 else ''})"
    prior_lbl = f"YTD {prior_year}" if has_prior else "No prior year data"

    return {
        "kpis":            kpis,
        "trend":           [],
        "revenue_split":   [],
        "expense_split":   [],
        "movements":       movements,
        "commentary":      commentary,
        "waterfall":       None,
        "period":          {"label": ytd_label, "prior": prior_lbl},
        "selected_period": str(sel_ts),
        "periods":         [str(p) for p in all_periods],
    }


# ─────────────────────────────────────────────
# INSIGHTS — margins, pareto, momentum, etc.
# ─────────────────────────────────────────────

_FIXED_COST_CATEGORIES = frozenset({
    "Staff Costs", "Management Costs", "Premises Costs", "IT Costs",
    "Professional Fees", "Insurance", "Depreciation & Amortisation",
    "Finance Costs", "Office & Admin", "Tax",
})
_VARIABLE_COST_CATEGORIES = frozenset({
    "Direct Costs", "Marketing Costs", "Travel & Entertainment",
})

_CATEGORY_ORDER = {
    "Revenue": 0, "Direct Costs": 1, "Staff Costs": 2, "Management Costs": 3,
    "IT Costs": 4, "Premises Costs": 5, "Professional Fees": 6,
    "Marketing Costs": 7, "Travel & Entertainment": 8, "Office & Admin": 9,
    "Insurance": 10, "Finance Costs": 11, "Depreciation & Amortisation": 12,
    "Tax": 13, "Other": 99,
}


def get_insights_data(
    analysis_df: pd.DataFrame,
    df_long: pd.DataFrame,
    selected_period,
    kpi_accounts: dict,
    mode: str,
) -> dict:
    """Compute supplementary FP&A insights for the Insights view."""
    sort_key = (lambda p: pd.Timestamp(p)) if mode == "monthly" else quarter_sort_key
    sorted_periods = sorted(analysis_df["Period"].unique(), key=sort_key)
    idx = (
        list(sorted_periods).index(selected_period)
        if selected_period in sorted_periods
        else len(sorted_periods) - 1
    )

    period_df = analysis_df[analysis_df["Period"] == selected_period].copy()
    driver_df = period_df[~period_df["Is Subtotal"]].copy()

    rev_name  = kpi_accounts.get("revenue") or ""
    cost_name = kpi_accounts.get("costs")   or ""
    prof_name = kpi_accounts.get("profit")  or ""

    def _kpi(name, df=None):
        if not name:
            return None
        d = period_df if df is None else df
        m = d[d["Account"].str.strip() == name.strip()]
        if m.empty:
            return None
        v = m.iloc[0]["Value"]
        return float(v) if pd.notna(v) else None

    rev   = _kpi(rev_name)
    costs = _kpi(cost_name)
    prof  = _kpi(prof_name)

    def _pct(num, denom):
        if num is None or denom is None or denom == 0:
            return None
        return round(num / denom * 100, 1)

    # ── 1. MARGINS ────────────────────────────────────────────────────
    staff_sum   = float(driver_df[driver_df["Category"] == "Staff Costs"]["Value"].sum())
    op_pct      = _pct(prof, rev)
    payroll_pct = _pct(staff_sum, rev)

    trend_margins = []
    for p in sorted_periods[max(0, idx - 5):idx + 1]:
        p_df  = analysis_df[analysis_df["Period"] == p]
        p_drv = p_df[~p_df["Is Subtotal"]]
        r, pr = _kpi(rev_name, p_df), _kpi(prof_name, p_df)
        sc    = float(p_drv[p_drv["Category"] == "Staff Costs"]["Value"].sum()) if not p_drv.empty else 0
        short = pd.Timestamp(p).strftime("%b") if mode == "monthly" else str(p)[-5:]
        trend_margins.append({
            "m": short, "full": period_label(p, mode),
            "op_pct": _pct(pr, r), "payroll_pct": _pct(sc, r),
            "revenue": r, "profit": pr,
        })

    margins = {
        "op_pct": op_pct, "payroll_pct": payroll_pct,
        "staff_value": staff_sum, "trend": trend_margins,
    }

    # ── 2. COMMON-SIZE P&L ────────────────────────────────────────────
    common_size = []
    if rev and rev != 0:
        rev_row   = period_df[period_df["Account"].str.strip() == rev_name.strip()]
        prior_rev = (
            float(rev_row.iloc[0]["Prior Value"])
            if not rev_row.empty and pd.notna(rev_row.iloc[0]["Prior Value"])
            else None
        )
        cs_df = driver_df.copy()
        cs_df["_ord"] = cs_df["Category"].map(lambda c: _CATEGORY_ORDER.get(c, 50))
        for _, row in cs_df.sort_values(["_ord", "Account"]).iterrows():
            v   = float(row["Value"])       if pd.notna(row["Value"])       else None
            pv  = float(row["Prior Value"]) if pd.notna(row["Prior Value"]) else None
            pct = _pct(v, rev)
            ppct = _pct(pv, prior_rev) if pv is not None and prior_rev else None
            common_size.append({
                "account": str(row["Account"]),
                "category": str(row["Category"]),
                "value": v,
                "pct_of_rev": pct,
                "prior_pct_of_rev": ppct,
                "delta_pp": round(pct - ppct, 1) if pct is not None and ppct is not None else None,
            })

    # ── 3. ROLLING 12 MONTHS (R12) ────────────────────────────────────
    r12 = {"available": False}
    if mode == "monthly" and idx + 1 >= 12:
        r12_ps = sorted_periods[idx - 11:idx + 1]

        def _sum(name, ps):
            if not name:
                return None
            return sum(
                v for p in ps
                for v in [_kpi(name, analysis_df[analysis_df["Period"] == p])]
                if v is not None
            )

        r12_rev, r12_cost, r12_prof = _sum(rev_name, r12_ps), _sum(cost_name, r12_ps), _sum(prof_name, r12_ps)
        r12 = {
            "available": True, "periods_used": len(r12_ps),
            "revenue": r12_rev, "costs": r12_cost, "profit": r12_prof,
            "op_pct": _pct(r12_prof, r12_rev),
        }
        if idx >= 23:
            pr_ps = sorted_periods[idx - 23:idx - 11]
            r12["prior_revenue"] = _sum(rev_name,  pr_ps)
            r12["prior_costs"]   = _sum(cost_name, pr_ps)
            r12["prior_profit"]  = _sum(prof_name, pr_ps)

    # ── 4. RUN-RATE ───────────────────────────────────────────────────
    run_rate = {
        "revenue": rev  * 12 if rev   is not None else None,
        "costs":   costs * 12 if costs is not None else None,
        "profit":  prof  * 12 if prof  is not None else None,
        "op_pct":  op_pct,
    }

    # ── 5. PARETO — cost concentration ───────────────────────────────
    cost_df = driver_df[driver_df["Category"] != "Revenue"].copy()
    cost_df["_abs"] = cost_df["Value"].abs()
    cost_df = cost_df[cost_df["_abs"] > 0].sort_values("_abs", ascending=False).reset_index(drop=True)
    total_cost = float(cost_df["_abs"].sum())
    cum, pareto_lines = 0.0, []
    for _, row in cost_df.iterrows():
        cum += float(row["_abs"])
        pareto_lines.append({
            "account": str(row["Account"]),
            "category": str(row["Category"]),
            "value": float(row["_abs"]),
            "pct_of_total": round(row["_abs"] / total_cost * 100, 1) if total_cost else 0,
            "cum_pct": round(cum / total_cost * 100, 1) if total_cost else 0,
        })
    top3_pct = pareto_lines[2]["cum_pct"] if len(pareto_lines) >= 3 else (pareto_lines[-1]["cum_pct"] if pareto_lines else 0)
    top5_pct = pareto_lines[4]["cum_pct"] if len(pareto_lines) >= 5 else (pareto_lines[-1]["cum_pct"] if pareto_lines else 0)
    pareto = {"lines": pareto_lines[:12], "top3_pct": top3_pct, "top5_pct": top5_pct, "total_cost": total_cost}

    # ── 6. SAME-PERIOD PRIOR YEAR (SPPY) ─────────────────────────────
    sppy = {"available": False}
    if mode == "monthly":
        sel_ts = pd.Timestamp(selected_period)
        sp_period = next(
            (p for p in sorted_periods
             if pd.Timestamp(p).year == sel_ts.year - 1 and pd.Timestamp(p).month == sel_ts.month),
            None,
        )
        if sp_period is not None:
            sp_df = analysis_df[analysis_df["Period"] == sp_period]
            sp_rev, sp_cost, sp_prof = _kpi(rev_name, sp_df), _kpi(cost_name, sp_df), _kpi(prof_name, sp_df)
            sppy = {
                "available": True,
                "period_label": period_label(sp_period, mode),
                "revenue": sp_rev, "costs": sp_cost, "profit": sp_prof,
                "rev_delta":  (rev   - sp_rev)  if rev   is not None and sp_rev  is not None else None,
                "cost_delta": (costs - sp_cost) if costs is not None and sp_cost is not None else None,
                "prof_delta": (prof  - sp_prof) if prof  is not None and sp_prof is not None else None,
                "rev_pct":    _pct(rev   - sp_rev,  sp_rev)  if rev   is not None and sp_rev  else None,
                "prof_pct":   _pct(prof  - sp_prof, sp_prof) if prof  is not None and sp_prof else None,
                "sp_op_pct":  _pct(sp_prof, sp_rev),
            }

    # ── 7. MOMENTUM ───────────────────────────────────────────────────
    momentum = {"available": False}
    if idx >= 5:
        last3  = sorted_periods[idx - 2:idx + 1]
        prior3 = sorted_periods[idx - 5:idx - 2]

        def _avg(name, ps):
            vals = [_kpi(name, analysis_df[analysis_df["Period"] == p]) for p in ps]
            vals = [v for v in vals if v is not None]
            return sum(vals) / len(vals) if vals else None

        def _dir(curr, prev, fav_up=True):
            if curr is None or prev is None or prev == 0:
                return "stable"
            chg = (curr - prev) / abs(prev)
            if chg > 0.02:
                return "improving" if fav_up else "worsening"
            if chg < -0.02:
                return "worsening" if fav_up else "improving"
            return "stable"

        rev_dir  = _dir(_avg(rev_name,  last3), _avg(rev_name,  prior3), True)
        prof_dir = _dir(_avg(prof_name, last3), _avg(prof_name, prior3), True)
        cost_dir = _dir(_avg(cost_name, last3), _avg(cost_name, prior3), False)
        dirs     = [rev_dir, prof_dir, cost_dir]
        overall  = (
            "improving" if dirs.count("improving") > dirs.count("worsening")
            else "worsening" if dirs.count("worsening") > dirs.count("improving")
            else "stable"
        )
        momentum = {
            "available": True, "overall": overall,
            "revenue_dir": rev_dir, "profit_dir": prof_dir, "cost_dir": cost_dir,
            "periods_compared": 3,
        }

    # ── 8. FIXED vs VARIABLE ──────────────────────────────────────────
    fixed_sum = float(driver_df[driver_df["Category"].isin(_FIXED_COST_CATEGORIES)]["Value"].sum())
    var_sum   = float(driver_df[driver_df["Category"].isin(_VARIABLE_COST_CATEGORIES)]["Value"].sum())
    contrib_pct, breakeven = None, None
    if rev and rev != 0 and abs(var_sum) < abs(rev):
        var_ratio   = abs(var_sum) / abs(rev)
        contrib_pct = round((1 - var_ratio) * 100, 1)
        if contrib_pct > 0 and fixed_sum:
            breakeven = abs(fixed_sum) / (1 - var_ratio)

    fixed_variable = {
        "fixed_cost": abs(fixed_sum), "variable_cost": abs(var_sum),
        "total_classified": abs(fixed_sum) + abs(var_sum),
        "contribution_margin_pct": contrib_pct,
        "breakeven_revenue": float(breakeven) if breakeven is not None else None,
        "has_revenue": rev is not None and rev != 0,
    }

    # ── 9. NON-RECURRING LINES ───────────────────────────────────────
    nonrecurring = []
    check_n = min(6, len(sorted_periods))
    if check_n >= 3:
        check_ps = set(sorted_periods[max(0, idx - check_n + 1):idx + 1])
        for acc in analysis_df[~analysis_df["Is Subtotal"]]["Account"].unique():
            acc_rows = analysis_df[(analysis_df["Account"] == acc) & (analysis_df["Period"].isin(check_ps))]
            present  = int(acc_rows[acc_rows["Value"].abs() > 0]["Period"].nunique())
            if 0 < present < len(check_ps) * 2 / 3:
                cat_m = driver_df[driver_df["Account"] == acc]["Category"]
                val_m = driver_df[driver_df["Account"] == acc]["Value"]
                cat   = str(cat_m.iloc[0]) if not cat_m.empty else "Other"
                val   = float(val_m.iloc[0]) if not val_m.empty else None
                nonrecurring.append({
                    "account": str(acc), "category": cat,
                    "present_count": present, "check_periods": len(check_ps),
                    "value": val, "in_current": val is not None and val != 0,
                })
        nonrecurring.sort(key=lambda x: (not x["in_current"], x["account"]))

    return {
        "margins":        margins,
        "common_size":    common_size,
        "r12":            r12,
        "run_rate":       run_rate,
        "pareto":         pareto,
        "sppy":           sppy,
        "momentum":       momentum,
        "fixed_variable": fixed_variable,
        "nonrecurring":   nonrecurring[:20],
        "period_label":   period_label(selected_period, mode),
    }


# ─────────────────────────────────────────────
# BUDGET VS ACTUAL — SHEET & COLUMN DETECTION
# ─────────────────────────────────────────────
BVA_ACTUAL_KEYWORDS     = ["actual", "actuals", "actual amount", "current actual", "ytd actual"]
BVA_BUDGET_KEYWORDS     = ["budget", "budget amount", "original budget", "revised budget", "forecast", "plan"]
BVA_ACTUAL_SHEET_NAMES  = ["actuals", "actual"]
BVA_BUDGET_SHEET_NAMES  = ["budget", "budgets"]

# Common FP&A column header date formats, tried in order
_BVA_DATE_FORMATS = [
    "%b-%y",   # Apr-24
    "%b-%Y",   # Apr-2024
    "%B-%y",   # April-24
    "%B-%Y",   # April-2024
    "%b %y",   # Apr 24
    "%b %Y",   # Apr 2024
    "%B %Y",   # April 2024
    "%Y-%m",   # 2024-04
    "%m/%Y",   # 04/2024
    "%m-%Y",   # 04-2024
]


def _try_parse_date_col(col_str: str):
    """
    Try to parse a column header string as a date/month using common FP&A formats.
    Handles bare 3-letter month names (Apr, May, Jun…) by assigning year 2000
    so they produce a sortable Timestamp even without an explicit year.
    Returns pd.NaT if nothing matches.
    """
    s = col_str.replace("Sept", "Sep").strip()
    # Bare 3- or 4-letter month abbreviation with no year (e.g. "Apr", "Sept")
    try:
        ts = pd.to_datetime(s + " 2000", format="%b %Y")
        return ts
    except (ValueError, TypeError):
        pass
    for fmt in _BVA_DATE_FORMATS:
        try:
            return pd.to_datetime(s, format=fmt)
        except (ValueError, TypeError):
            pass
    # generic fallback (handles ISO dates, 'Apr 2024', etc.)
    return pd.to_datetime(s, errors="coerce")


def _parse_wide_sheet(raw: pd.DataFrame) -> tuple:
    """
    Parse a raw (header=None) DataFrame — as returned by pd.read_excel(..., header=None) —
    into (df_with_Account_Section_plus_month_columns, list_of_month_col_names).

    Handles both:
      • Two-column layout: Section/Level_1 col | Account col | month cols…
      • Single-column layout: Account col | month cols…

    Raises ValueError if no Account header row is found.
    """
    header_row = None
    for i in range(min(25, len(raw))):
        if "account" in raw.iloc[i].astype(str).str.strip().str.lower().tolist():
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find an 'Account' header row in sheet.")

    headers = raw.iloc[header_row].astype(str).str.strip().tolist()
    df = raw.iloc[header_row + 1:].copy()
    df.columns = headers

    first_col   = df.columns[0]
    account_col = next((c for c in df.columns if c.strip().lower() == "account"), None)
    if account_col is None:
        raise ValueError("Could not find the Account column.")

    rename_map = {}
    if first_col != account_col:
        rename_map[first_col] = "Level_1"
    rename_map[account_col] = "Level_2_Account"
    df = df.rename(columns=rename_map)
    if "Level_1" not in df.columns:
        df["Level_1"] = ""

    reserved    = {"Level_1", "Level_2_Account"}
    month_cols  = []
    for col in df.columns:
        if col in reserved:
            continue
        parsed = _try_parse_date_col(str(col))
        if pd.notna(parsed):
            month_cols.append(col)

    section = None
    rows    = []
    for _, row in df.iterrows():
        left = clean_text(row.get("Level_1", ""))
        acc  = clean_text(row.get("Level_2_Account", ""))

        if is_heading_only_row(left, acc):
            section = left
            continue

        account_name = first_non_blank([acc, left])
        if not account_name:
            continue

        r = {"Account": account_name, "Section": section or ""}
        for col in month_cols:
            r[col] = row.get(col)
        rows.append(r)

    return pd.DataFrame(rows), month_cols


def load_bva_from_sheets(contents: bytes, filename: str) -> tuple:
    """
    Detect and load Actuals and Budget sheets from an Excel workbook.

    Sheet name matching is case-insensitive.
    Acceptable names:
        Actuals: "Actuals", "Actual"
        Budget : "Budget",  "Budgets"

    Returns (df_actual, df_budget) — both wide DataFrames with Account, Section, month cols.
    Raises ValueError with a descriptive message if sheets are not found.

    Logs to stderr:
        • All sheet names found
        • Actual sheet selected
        • Budget sheet selected
        • Account counts per sheet
        • Month/period counts per sheet
    """
    xf          = pd.ExcelFile(BytesIO(contents))
    all_sheets  = xf.sheet_names
    sheets_lower = {s.lower().strip(): s for s in all_sheets}

    print(f"[BvA] Sheets found    : {all_sheets}", file=sys.stderr)

    actual_sheet = next(
        (sheets_lower[k] for k in BVA_ACTUAL_SHEET_NAMES if k in sheets_lower), None
    )
    budget_sheet = next(
        (sheets_lower[k] for k in BVA_BUDGET_SHEET_NAMES if k in sheets_lower), None
    )

    if not actual_sheet:
        raise ValueError(
            f"No Actuals sheet found. Available sheets: {all_sheets}. "
            f"Expected one of: {BVA_ACTUAL_SHEET_NAMES}"
        )
    if not budget_sheet:
        raise ValueError(
            f"No Budget sheet found. Available sheets: {all_sheets}. "
            f"Expected one of: {BVA_BUDGET_SHEET_NAMES}"
        )

    print(f"[BvA] Actual sheet    : '{actual_sheet}'", file=sys.stderr)
    print(f"[BvA] Budget sheet    : '{budget_sheet}'", file=sys.stderr)

    raw_actual = pd.read_excel(BytesIO(contents), sheet_name=actual_sheet, header=None)
    raw_budget = pd.read_excel(BytesIO(contents), sheet_name=budget_sheet, header=None)

    df_actual, actual_month_cols = _parse_wide_sheet(raw_actual)
    df_budget, budget_month_cols = _parse_wide_sheet(raw_budget)

    print(f"[BvA] Actual accounts : {len(df_actual)}, periods: {len(actual_month_cols)}", file=sys.stderr)
    print(f"[BvA] Budget accounts : {len(df_budget)}, periods: {len(budget_month_cols)}", file=sys.stderr)

    return df_actual, df_budget


def build_bva_long_from_sheets(df_actual: pd.DataFrame, df_budget: pd.DataFrame) -> pd.DataFrame:
    """
    Merge Actuals and Budget wide DataFrames by Account + matched monthly columns.

    Returns a long-form DataFrame with columns:
        Account, Section, Category, Is Subtotal,
        Period (Timestamp), Actual, Budget, Variance, Variance %, Abs Variance

    Month columns are matched by normalising both sides to pd.Timestamp so that
    different string representations (e.g. "Apr-24" vs "2024-04-01") still align.
    """
    def month_ts_map(df):
        """Return {Timestamp: original_col_name} for every parseable date column."""
        result = {}
        for col in df.columns:
            if col in ("Account", "Section"):
                continue
            ts = _try_parse_date_col(str(col))
            if pd.notna(ts):
                result[pd.Timestamp(ts)] = col
        return result

    actual_months = month_ts_map(df_actual)
    budget_months = month_ts_map(df_budget)
    common_ts     = sorted(set(actual_months.keys()) & set(budget_months.keys()))

    print(
        f"[BvA] Actual periods  : {[pd.Timestamp(t).strftime('%b %Y') for t in sorted(actual_months)]}",
        file=sys.stderr,
    )
    print(
        f"[BvA] Budget periods  : {[pd.Timestamp(t).strftime('%b %Y') for t in sorted(budget_months)]}",
        file=sys.stderr,
    )
    print(
        f"[BvA] Common periods  : {[pd.Timestamp(t).strftime('%b %Y') for t in common_ts]}",
        file=sys.stderr,
    )

    if not common_ts:
        raise ValueError(
            "No common month columns found between the Actuals and Budget sheets. "
            "Ensure both sheets use matching column headers (e.g. 'Apr-24' on both)."
        )

    actual_accounts = set(df_actual["Account"].dropna().str.strip().str.lower())
    budget_accounts = set(df_budget["Account"].dropna().str.strip().str.lower())
    matched_count   = len(actual_accounts & budget_accounts)
    print(
        f"[BvA] Accounts — Actual: {len(actual_accounts)}, "
        f"Budget: {len(budget_accounts)}, Matched: {matched_count}",
        file=sys.stderr,
    )

    rows = []
    for ts in common_ts:
        a_col = actual_months[ts]
        b_col = budget_months[ts]

        df_a = df_actual[["Account", "Section", a_col]].copy().rename(columns={a_col: "Actual"})
        df_b = df_budget[["Account",             b_col]].copy().rename(columns={b_col: "Budget"})

        df_a["Actual"] = clean_numeric(df_a["Actual"])
        df_b["Budget"] = clean_numeric(df_b["Budget"])

        merged         = df_a.merge(df_b, on="Account", how="outer")
        merged["Period"] = ts
        rows.append(merged)

    df_long = pd.concat(rows, ignore_index=True)
    df_long["Section"]    = df_long["Section"].fillna("").astype(str)
    df_long["Actual"]     = df_long["Actual"].fillna(0)
    df_long["Budget"]     = df_long["Budget"].fillna(0)
    df_long["Variance"]   = df_long["Actual"] - df_long["Budget"]
    df_long["Variance %"] = df_long.apply(
        lambda r: (r["Variance"] / r["Budget"] * 100) if r["Budget"] != 0 else None, axis=1
    )
    df_long["Abs Variance"] = df_long["Variance"].abs()
    df_long["Category"]     = df_long.apply(lambda r: classify(r["Account"], r["Section"]), axis=1)
    df_long["Is Subtotal"]  = df_long["Account"].apply(is_subtotal)
    return df_long


def detect_bva_columns(df: pd.DataFrame) -> tuple:
    """
    Return (actual_col, budget_col) by keyword-matching df.columns.
    Tries exact match first, then partial match.
    Returns (None, None) if not found.
    """
    cols_lower = {c.lower().strip(): c for c in df.columns}

    def find_col(keywords):
        # Exact match
        for kw in keywords:
            if kw in cols_lower:
                return cols_lower[kw]
        # Partial match
        for kw in keywords:
            for c_lower, c in cols_lower.items():
                if kw in c_lower:
                    return c
        return None

    return find_col(BVA_ACTUAL_KEYWORDS), find_col(BVA_BUDGET_KEYWORDS)


# ─────────────────────────────────────────────
# BUDGET VS ACTUAL — CORE BUILD
# ─────────────────────────────────────────────
def build_bva(df: pd.DataFrame, actual_col: str, budget_col: str) -> pd.DataFrame:
    """
    Build a Budget vs Actual analysis dataframe.

    Output columns:
        Account, Section, Category, Is Subtotal,
        Actual, Budget, Variance (Actual − Budget), Variance %, Abs Variance
    """
    bva = df[["Account", "Section", actual_col, budget_col]].copy()
    bva = bva.rename(columns={actual_col: "Actual", budget_col: "Budget"})
    bva["Actual"] = clean_numeric(bva["Actual"])
    bva["Budget"] = clean_numeric(bva["Budget"])
    bva = bva.dropna(subset=["Actual", "Budget"], how="all")
    bva["Actual"] = bva["Actual"].fillna(0)
    bva["Budget"] = bva["Budget"].fillna(0)
    bva["Variance"] = bva["Actual"] - bva["Budget"]
    bva["Variance %"] = bva.apply(
        lambda r: (r["Variance"] / r["Budget"] * 100) if r["Budget"] != 0 else None, axis=1
    )
    bva["Abs Variance"] = bva["Variance"].abs()
    bva["Category"]    = bva.apply(lambda r: classify(r["Account"], r["Section"]), axis=1)
    bva["Is Subtotal"] = bva["Account"].apply(is_subtotal)
    return bva


# ─────────────────────────────────────────────
# BUDGET VS ACTUAL — WATERFALL
# ─────────────────────────────────────────────
def build_bva_waterfall(df_bva: pd.DataFrame, kpi_accounts: dict) -> dict | None:
    """
    Compute what drove the difference between Budget and Actual Operating Profit.

    Profit impact formula:
        Revenue: actual − budget   (actual > budget = favourable = positive impact)
        Costs:   budget − actual   (actual < budget = favourable = positive impact)
    """
    profit_account = kpi_accounts.get("profit")
    if not profit_account:
        return None

    prof_row = df_bva[df_bva["Account"].str.strip() == profit_account.strip()]
    if prof_row.empty:
        return None

    r = prof_row.iloc[0]
    budget_profit = float(r["Budget"]) if pd.notna(r["Budget"]) else 0.0
    actual_profit = float(r["Actual"]) if pd.notna(r["Actual"]) else 0.0
    net_change    = actual_profit - budget_profit

    def _is_revenue(row) -> bool:
        return row.get("Category") == "Revenue"

    def _profit_impact(row) -> float:
        actual = float(row["Actual"]) if pd.notna(row.get("Actual")) else 0.0
        budget = float(row["Budget"]) if pd.notna(row.get("Budget")) else 0.0
        return (actual - budget) if _is_revenue(row) else (budget - actual)

    driver_df = df_bva[~df_bva["Is Subtotal"]].copy()
    driver_df["ProfitImpact"] = driver_df.apply(_profit_impact, axis=1)

    cat_agg = (
        driver_df.groupby("Category")
        .agg(
            impact       =("ProfitImpact", "sum"),
            actual_total =("Actual", lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum()),
            budget_total =("Budget", lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum()),
        )
        .reset_index()
    )

    category_detail = [
        {
            "category"    : str(row["Category"]),
            "prior"       : round(float(row["budget_total"]), 2),   # Budget = "prior" for display compatibility
            "current"     : round(float(row["actual_total"]), 2),   # Actual = "current"
            "raw_variance": round(float(row["actual_total"]) - float(row["budget_total"]), 2),
            "profit_impact": round(float(row["impact"]), 2),
            "is_revenue"  : _is_revenue({"Category": str(row["Category"])}),
        }
        for _, row in cat_agg.iterrows()
    ]

    cat_df = cat_agg[["Category", "impact"]].copy()
    cat_df = cat_df[cat_df["impact"].abs().round(2) > 0]

    threshold = abs(net_change) * 0.05 if abs(net_change) > 0 else 0
    if threshold > 0:
        material   = cat_df[cat_df["impact"].abs() >= threshold].copy()
        immaterial = cat_df[cat_df["impact"].abs() <  threshold].copy()
    else:
        material   = cat_df.copy()
        immaterial = pd.DataFrame(columns=cat_df.columns)

    if not immaterial.empty:
        other_sum = float(immaterial["impact"].sum())
        if abs(other_sum) > 0.01:
            material = pd.concat(
                [material, pd.DataFrame([{"Category": "Other Movements", "impact": other_sum}])],
                ignore_index=True,
            )

    def _sort_key(row):
        cat = row["Category"]
        if cat == "Revenue":         return (0, -abs(row["impact"]))
        if cat == "Other Movements": return (99, 0)
        return                              (1,  -abs(row["impact"]))

    material["_sk"] = material.apply(_sort_key, axis=1)
    material = material.sort_values("_sk").drop(columns=["_sk"])

    bars = [
        {"label": r["Category"], "impact": round(float(r["impact"]), 2), "fav": float(r["impact"]) > 0}
        for _, r in material.iterrows()
    ]

    # Reconciliation
    explained  = round(sum(b["impact"] for b in bars), 2)
    residual   = round(net_change - explained, 2)
    reconciles = abs(residual) <= 0.5
    if not reconciles:
        other_idx = next((i for i, b in enumerate(bars) if b["label"] == "Other Movements"), None)
        if other_idx is not None:
            merged = round(bars[other_idx]["impact"] + residual, 2)
            bars[other_idx] = {"label": "Other Movements", "impact": merged, "fav": merged > 0}
        else:
            bars.append({"label": "Other Movements", "impact": residual, "fav": residual > 0})

    non_other   = [b for b in bars if b["label"] != "Other Movements"]
    pos_bars    = [b for b in non_other if b["impact"] > 0]
    neg_bars    = [b for b in non_other if b["impact"] < 0]
    largest_pos = max(pos_bars, key=lambda b: b["impact"])  if pos_bars else None
    largest_neg = min(neg_bars, key=lambda b: b["impact"])  if neg_bars else None

    # Commentary — Budget vs Actual wording
    direction = "above" if net_change >= 0 else "below"
    fav_word  = "favourable" if net_change >= 0 else "adverse"
    parts = [f"{profit_account} was {fmt_gbp(abs(net_change))} {direction} budget ({fav_word})."]

    if largest_pos:
        if largest_pos["label"] == "Revenue":
            parts.append(f"Revenue exceeded budget by {fmt_gbp(largest_pos['impact'])}, contributing positively to profit.")
        else:
            parts.append(f"{largest_pos['label']} came in under budget by {fmt_gbp(abs(largest_pos['impact']))}, improving profit.")
    if largest_neg:
        if largest_neg["label"] == "Revenue":
            parts.append(f"Revenue fell short of budget by {fmt_gbp(abs(largest_neg['impact']))}, reducing profit vs budget.")
        else:
            parts.append(f"{largest_neg['label']} exceeded budget by {fmt_gbp(abs(largest_neg['impact']))}, reducing profit.")

    return {
        "prior_profit"   : round(budget_profit, 2),
        "current_profit" : round(actual_profit, 2),
        "net_change"     : round(net_change, 2),
        "bars"           : bars,
        "largest_positive": largest_pos,
        "largest_negative": largest_neg,
        "commentary"     : " ".join(parts),
        "profit_account" : profit_account,
        "reconciles"     : reconciles,
        "category_detail": category_detail,
    }


# ─────────────────────────────────────────────
# BUDGET VS ACTUAL — TREND BUILDER
# ─────────────────────────────────────────────
def build_bva_trend(bva_long: pd.DataFrame, kpi_accounts: dict) -> list:
    if bva_long is None or bva_long.empty:
        return []
    periods = sorted(bva_long["Period"].unique())
    if len(periods) < 2:
        return []
    rev_name  = kpi_accounts.get("revenue")
    cost_name = kpi_accounts.get("costs")
    prof_name = kpi_accounts.get("profit")
    trend = []
    for ts in periods:
        pdf = bva_long[bva_long["Period"] == ts]
        short = pd.Timestamp(ts).strftime("%b")
        full  = pd.Timestamp(ts).strftime("%B %Y")
        def _val(account_name, col):
            if not account_name:
                return 0
            m = pdf[pdf["Account"].str.strip() == account_name.strip()]
            if m.empty:
                return 0
            v = m.iloc[0].get(col)
            return float(v) if pd.notna(v) else 0
        trend.append({
            "m": short, "full": full,
            "actual_revenue": _val(rev_name, "Actual"),
            "budget_revenue": _val(rev_name, "Budget"),
            "actual_costs":   _val(cost_name, "Actual"),
            "budget_costs":   _val(cost_name, "Budget"),
            "actual_profit":  _val(prof_name, "Actual"),
            "budget_profit":  _val(prof_name, "Budget"),
        })
    return trend


# BUDGET VS ACTUAL — PERIOD DATA FOR API
# ─────────────────────────────────────────────
def get_bva_data(df_bva: pd.DataFrame, kpi_accounts: dict, filename: str, bva_long: pd.DataFrame | None = None) -> dict:
    """Return the full dashboard data dict for Budget vs Actual mode."""
    driver_df = df_bva[~df_bva["Is Subtotal"]].copy()

    # ── KPI cards ──────────────────────────────────────────────────────────
    def kpi_row_bva(name, fav_when_actual_up, icon):
        if not name:
            return None
        match = df_bva[df_bva["Account"].str.strip() == name.strip()]
        if match.empty:
            return None
        r        = match.iloc[0]
        actual   = float(r["Actual"])     if pd.notna(r["Actual"])     else None
        budget   = float(r["Budget"])     if pd.notna(r["Budget"])     else None
        variance = float(r["Variance"])   if pd.notna(r["Variance"])   else None
        pct      = float(r["Variance %"]) if pd.notna(r["Variance %"]) else None
        is_fav   = (
            (fav_when_actual_up and variance >= 0) or
            (not fav_when_actual_up and variance <= 0)
        ) if variance is not None else None
        return {"label": name, "value": actual, "variance": variance,
                "prior": budget, "pct": pct, "is_fav": is_fav, "icon": icon, "pct_only": False}

    rev_kpi  = kpi_row_bva(kpi_accounts.get("revenue"), True,  "trending-up")
    cost_kpi = kpi_row_bva(kpi_accounts.get("costs"),   False, "receipt")
    prof_kpi = kpi_row_bva(kpi_accounts.get("profit"),  True,  "wallet")

    # ── Classification-based fallback ─────────────────────────────────────
    # Used when no named total/subtotal row was matched by detect_kpis.
    # Sums driver_df (non-subtotal rows) by Category.
    # Costs are assumed positive (expenses stored as positive numbers).
    def _sum_col(df, col, category_filter):
        """Sum a column for rows matching category_filter (None = all rows)."""
        subset = df if category_filter is None else df[df["Category"] == category_filter]
        return float(subset[col].sum())

    if rev_kpi is None:
        rev_act = _sum_col(driver_df, "Actual", "Revenue")
        rev_bud = _sum_col(driver_df, "Budget", "Revenue")
        rev_var = rev_act - rev_bud
        rev_pct = (rev_var / rev_bud * 100) if rev_bud != 0 else None
        rev_kpi = {
            "label": "Total Revenue", "value": rev_act, "prior": rev_bud,
            "variance": rev_var, "pct": rev_pct, "is_fav": rev_var >= 0,
            "icon": "trending-up", "pct_only": False,
        }

    if cost_kpi is None:
        cost_act = _sum_col(driver_df[driver_df["Category"] != "Revenue"], "Actual", None)
        cost_bud = _sum_col(driver_df[driver_df["Category"] != "Revenue"], "Budget", None)
        cost_var = cost_act - cost_bud
        cost_pct = (cost_var / cost_bud * 100) if cost_bud != 0 else None
        cost_kpi = {
            "label": "Total Costs", "value": cost_act, "prior": cost_bud,
            "variance": cost_var, "pct": cost_pct, "is_fav": cost_var <= 0,
            "icon": "receipt", "pct_only": False,
        }

    if prof_kpi is None:
        rev_act  = rev_kpi["value"]
        rev_bud  = rev_kpi["prior"]
        cost_act = cost_kpi["value"]
        cost_bud = cost_kpi["prior"]
        prof_act = rev_act - cost_act
        prof_bud = rev_bud - cost_bud
        prof_var = prof_act - prof_bud
        prof_pct = (prof_var / abs(prof_bud) * 100) if prof_bud != 0 else None
        prof_kpi = {
            "label": "Operating Profit", "value": prof_act, "prior": prof_bud,
            "variance": prof_var, "pct": prof_pct, "is_fav": prof_var >= 0,
            "icon": "wallet", "pct_only": False,
        }

    kpis = [k for k in [rev_kpi, cost_kpi, prof_kpi] if k]
    if prof_kpi and prof_kpi["pct"] is not None:
        kpis.append({
            "label": "Profit vs Budget", "value": None, "variance": prof_kpi["variance"],
            "pct": prof_kpi["pct"], "is_fav": prof_kpi["is_fav"],
            "icon": "bar-chart-2", "pct_only": True,
        })

    # ── Revenue split (Actual values) ───────────────────────────────────────
    def revenue_split_fn(top_n=8):
        SECTION_HEADING_NAMES = {"turnover", "income", "revenue", "receipts"}
        SUBTOTAL_PATTERNS = ["total", "gross profit", "subtotal", "sub-total",
                             "net revenue", "net income", "total income"]

        def looks_like_subtotal(name):
            n = name.lower().strip()
            return n in SECTION_HEADING_NAMES or any(p in n for p in SUBTOTAL_PATTERNS)

        df = driver_df[
            (driver_df["Category"] == "Revenue") &
            (~driver_df["Account"].apply(looks_like_subtotal))
        ].copy()
        df = df.groupby("Account", as_index=False)["Actual"].sum()
        df["Chart Value"] = df["Actual"].abs()
        df = df[df["Chart Value"] > 0].sort_values("Chart Value", ascending=False)
        if df.empty:
            return []
        if len(df) > top_n:
            other = df.iloc[top_n:]["Chart Value"].sum()
            top   = df.head(top_n).copy()
            df = pd.concat(
                [top, pd.DataFrame([{"Account": "Other", "Actual": other, "Chart Value": other}])],
                ignore_index=True,
            ) if other > 0 else top
        total = df["Chart Value"].sum()
        return [{"name": r["Account"], "value": float(r["Chart Value"]),
                 "pct": round(r["Chart Value"] / total * 100, 1) if total else 0}
                for _, r in df.iterrows()]

    def expense_split_fn(categories, top_n=6):
        df = driver_df[driver_df["Category"].isin(categories)].copy()
        df = df.groupby("Category", as_index=False)["Actual"].sum()
        df["Chart Value"] = df["Actual"].abs()
        df = df[df["Chart Value"] > 0].sort_values("Chart Value", ascending=False)
        if len(df) > top_n:
            other = df.iloc[top_n:]["Chart Value"].sum()
            top   = df.head(top_n).copy()
            df = pd.concat(
                [top, pd.DataFrame([{"Category": "Other", "Actual": other, "Chart Value": other}])],
                ignore_index=True,
            ) if other > 0 else top
        total = df["Chart Value"].sum()
        return [{"name": r["Category"], "value": float(r["Chart Value"]),
                 "pct": round(r["Chart Value"] / total * 100, 1) if total else 0}
                for _, r in df.iterrows()]

    revenue_split = revenue_split_fn()
    expense_split = expense_split_fn(EXPENSE_CATEGORIES)

    # ── Top movements (by absolute variance) ────────────────────────────────
    top_mov = (
        driver_df[driver_df["Variance"].notna() & (driver_df["Variance"] != 0)]
        .sort_values("Abs Variance", ascending=False)
        .head(15)
    )
    movements = []
    for _, row in top_mov.iterrows():
        variance = float(row["Variance"]) if pd.notna(row["Variance"]) else None
        cat      = row["Category"]
        is_fav   = (
            (cat == "Revenue" and variance >= 0) or
            (cat != "Revenue" and variance <= 0)
        ) if variance is not None else None
        movements.append({
            "account"     : row["Account"],
            "category"    : cat,
            "value"       : float(row["Actual"])     if pd.notna(row["Actual"])     else None,
            "prior_value" : float(row["Budget"])     if pd.notna(row["Budget"])     else None,
            "variance"    : variance,
            "variance_pct": float(row["Variance %"]) if pd.notna(row["Variance %"]) else None,
            "is_fav"      : is_fav,
        })

    # ── Commentary ──────────────────────────────────────────────────────────
    commentary = []
    if prof_kpi and prof_kpi["variance"] is not None:
        word = "above" if prof_kpi["variance"] > 0 else "below"
        icon = "trending-up" if prof_kpi["variance"] > 0 else "trending-down"
        commentary.append({
            "icon": icon, "fav": prof_kpi["variance"] > 0,
            "html": (
                f"<b>{kpi_accounts.get('profit') or 'Operating profit'} was "
                f"{fmt_gbp(abs(prof_kpi['variance']))} {word} budget</b>"
                + (f" ({fmt_pct(abs(prof_kpi['pct']))})" if prof_kpi["pct"] else "") + "."
            ),
            "source": {"account": kpi_accounts.get('profit') or 'Operating profit',
                       "actual": prof_kpi["value"], "actual_label": "Actual",
                       "comparison": prof_kpi["prior"], "comparison_label": "Budget",
                       "variance": prof_kpi["variance"], "pct": prof_kpi["pct"]},
        })
    if rev_kpi and rev_kpi["variance"]:
        word = "above" if rev_kpi["variance"] > 0 else "below"
        commentary.append({
            "icon": "arrow-up-right" if rev_kpi["variance"] > 0 else "arrow-down-right",
            "fav" : rev_kpi["variance"] > 0,
            "html": (
                f"Revenue was <b>{fmt_gbp(abs(rev_kpi['variance']))} {word} budget</b>"
                + (f" ({fmt_pct(abs(rev_kpi['pct']))})" if rev_kpi["pct"] else "") + "."
            ),
            "source": {"account": rev_kpi["label"],
                       "actual": rev_kpi["value"], "actual_label": "Actual",
                       "comparison": rev_kpi["prior"], "comparison_label": "Budget",
                       "variance": rev_kpi["variance"], "pct": rev_kpi["pct"]},
        })
    for _, row in (driver_df[driver_df["Variance"].notna()]
                   .sort_values("Abs Variance", ascending=False).head(3).iterrows()):
        if row["Variance"] == 0:
            continue
        cat     = row["Category"]
        is_fav  = (cat == "Revenue" and row["Variance"] > 0) or (cat != "Revenue" and row["Variance"] < 0)
        word    = "above" if row["Variance"] > 0 else "below"
        pct_str = f" ({fmt_pct(abs(row['Variance %']))})" if pd.notna(row["Variance %"]) else ""
        commentary.append({
            "icon": "arrow-up-right" if row["Variance"] > 0 else "arrow-down-right",
            "fav" : is_fav,
            "html": f"<b>{row['Account']}</b> was {fmt_gbp(abs(row['Variance']))} {word} budget{pct_str}.",
            "source": {"account": row["Account"], "category": cat,
                       "actual": float(row["Actual"]) if pd.notna(row["Actual"]) else None, "actual_label": "Actual",
                       "comparison": float(row["Budget"]) if pd.notna(row["Budget"]) else None, "comparison_label": "Budget",
                       "variance": float(row["Variance"]), "pct": float(row["Variance %"]) if pd.notna(row["Variance %"]) else None},
        })

    waterfall = build_bva_waterfall(df_bva, kpi_accounts)

    return {
        "analysis_type" : "budget_vs_actual",
        "kpis"          : kpis,
        "trend"         : build_bva_trend(bva_long, kpi_accounts),
        "revenue_split" : revenue_split,
        "expense_split" : expense_split,
        "movements"     : movements,
        "commentary"    : commentary,
        "waterfall"     : waterfall,
        "period"        : {"label": "Actual", "prior": "Budget"},
        "selected_period": "budget_vs_actual",
        "periods"       : [],
        "file_name"     : filename,
    }


# ─────────────────────────────────────────────
# EXPORTS
# ─────────────────────────────────────────────
def make_zip(period_label_str: str, movements: list, commentary: list, kpis: list, currency_sym: str = "£") -> bytes:
    buf = BytesIO()
    safe = period_label_str.replace(" ", "_").replace("/", "-")
    _fg = lambda v: fmt_gbp(v, currency_sym)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Movements CSV
        rows = [",".join(["Account","Category","Value","Prior","Variance","Variance %"])]
        for m in movements:
            rows.append(",".join(str(x or "") for x in [
                m["account"], m["category"], m["value"], m["prior_value"],
                m["variance"], m["variance_pct"],
            ]))
        zf.writestr(f"variance_detail_{safe}.csv", "\n".join(rows))
        # Commentary text
        lines = [f"Management Pack — {period_label_str}", ""]
        for k in kpis:
            if not k.get("pct_only"):
                lines.append(f"{k['label']}: {_fg(k['value'])} (variance: {_fg(k['variance'])})")
        lines += ["", "Commentary:"] + [f"- {c['html']}" for c in commentary]
        zf.writestr(f"management_commentary_{safe}.txt", "\n".join(lines))
    buf.seek(0)
    return buf.getvalue()


def _fmt_signed_gbp(v, sym="£"):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return ("+" if v > 0 else "") + fmt_gbp(v, sym)


def _build_report_text(period_label_str, movements, commentary, kpis,
                       analysis_type="month_on_month", waterfall=None, currency_sym="£"):
    """Build structured report text blocks reusable by any export format."""
    import html as html_lib
    is_bva = analysis_type == "budget_vs_actual"
    prior_word = "Budget" if is_bva else "Prior"

    _fg  = lambda v: fmt_gbp(v, currency_sym)
    _fgs = lambda v: _fmt_signed_gbp(v, currency_sym)

    prof_kpi = next((k for k in kpis if not k.get("pct_only") and k.get("icon") == "wallet"), None)
    rev_kpi  = next((k for k in kpis if not k.get("pct_only") and k.get("icon") == "trending-up"), None)

    fav_rows = [m for m in movements if m.get("is_fav") and m.get("variance", 0) != 0]
    adv_rows = [m for m in movements if not m.get("is_fav") and m.get("variance", 0) != 0]
    fav_rows.sort(key=lambda m: abs(m.get("variance", 0)), reverse=True)
    adv_rows.sort(key=lambda m: abs(m.get("variance", 0)), reverse=True)
    largest_fav = fav_rows[0] if fav_rows else None
    largest_adv = adv_rows[0] if adv_rows else None
    total_fav = sum(abs(m.get("variance", 0)) for m in fav_rows)
    total_adv = sum(abs(m.get("variance", 0)) for m in adv_rows)

    # Executive summary
    exec_lines = []
    if prof_kpi:
        pv = prof_kpi.get("variance", 0)
        if is_bva:
            exec_lines.append(
                f"Operating Profit was {_fg(abs(pv))} {'above' if pv >= 0 else 'below'} budget "
                f"at {_fg(prof_kpi['value'])}, versus a budget of {_fg(prof_kpi['prior'])}."
            )
        else:
            exec_lines.append(
                f"Operating Profit {'increased' if pv >= 0 else 'decreased'} by {_fg(abs(pv))} "
                f"to {_fg(prof_kpi['value'])}, from {_fg(prof_kpi['prior'])} in the prior period."
            )
    if rev_kpi:
        exec_lines.append(
            f"Total Revenue was {_fg(rev_kpi['value'])} against "
            f"{'a budget' if is_bva else 'prior'} of {_fg(rev_kpi['prior'])} "
            f"({fmt_pct(rev_kpi.get('pct'))}%)." if rev_kpi.get("pct") is not None
            else f"Total Revenue was {_fg(rev_kpi['value'])}."
        )
    exec_lines.append(
        f"There were {len(fav_rows)} favourable variances totalling {_fg(total_fav)} "
        f"and {len(adv_rows)} adverse variances totalling {_fg(total_adv)}."
    )
    if largest_fav:
        exec_lines.append(f"Largest favourable: {largest_fav['account']} ({_fgs(largest_fav['variance'])})")
    if largest_adv:
        exec_lines.append(f"Largest adverse: {largest_adv['account']} ({_fgs(largest_adv['variance'])})")

    # Commentary (clean HTML)
    commentary_lines = []
    for c in commentary:
        clean = html_lib.unescape(re.sub("<[^>]+>", "", c.get("html", "")))
        if clean.strip():
            commentary_lines.append(clean.strip())

    # Recommended actions
    actions = []
    if largest_adv:
        pct_str = f" ({abs(largest_adv.get('variance_pct', 0)):.1f}%)" if largest_adv.get("variance_pct") else ""
        actions.append(
            f"[HIGH] Investigate {largest_adv['account']} — largest adverse variance at "
            f"{_fg(abs(largest_adv['variance']))}{pct_str}."
        )
    for m in adv_rows[1:3]:
        actions.append(
            f"[MEDIUM] Review {m['account']} ({m['category']}) — adverse variance of "
            f"{_fg(abs(m['variance']))}."
        )
    if largest_fav:
        actions.append(
            f"[LOW] Validate {largest_fav['account']} favourable variance of "
            f"{_fg(abs(largest_fav['variance']))} — confirm sustainability."
        )
    if prof_kpi and not prof_kpi.get("is_fav"):
        actions.append(
            f"[HIGH] Operating Profit is {_fg(abs(prof_kpi['variance']))} "
            f"{'below budget' if is_bva else 'below prior period'}. Prepare remediation plan."
        )

    return {
        "executive_summary": exec_lines,
        "commentary": commentary_lines,
        "key_findings_fav": fav_rows[:5],
        "key_findings_adv": adv_rows[:5],
        "actions": actions,
        "prior_word": prior_word,
    }


def make_pdf(period_label_str: str, movements: list, commentary: list, kpis: list,
             analysis_type: str = "month_on_month", waterfall: dict | None = None,
             firm_name: str = "", currency_sym: str = "£",
             insights: dict | None = None) -> bytes:
    """Generate a professional management pack PDF using ReportLab platypus."""
    import datetime
    from reportlab.lib import colors as C
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame, NextPageTemplate,
        Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether,
    )
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT

    # ── Brand colours ────────────────────────────────────────────────────────
    NAVY    = C.HexColor("#0C1726")
    PRIMARY = C.HexColor("#2F62E8")
    FAV_C   = C.HexColor("#0E8A57")
    ADV_C   = C.HexColor("#D02B45")
    FAV_BG  = C.HexColor("#E6F4EE")
    ADV_BG  = C.HexColor("#FDE8EC")
    SOFT    = C.HexColor("#F7F9FC")
    BORDER  = C.HexColor("#E2E8F0")
    FG2     = C.HexColor("#344054")
    FG3     = C.HexColor("#667085")
    AMBER   = C.HexColor("#D97706")
    AMBER_BG = C.HexColor("#FEF3C7")

    is_bva     = analysis_type == "budget_vs_actual"
    prior_word  = "Budget"  if is_bva else "Prior"
    actual_word = "Actual"  if is_bva else "Current"
    report = _build_report_text(period_label_str, movements, commentary, kpis, analysis_type, waterfall, currency_sym)
    gen_date = datetime.datetime.now().strftime("%d %B %Y")

    # ── Formatters ───────────────────────────────────────────────────────────
    _sym = currency_sym.strip()

    def _f(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "—"
        try:
            v = float(v)
        except (TypeError, ValueError):
            return "—"
        return (f"-{_sym}" if v < 0 else _sym) + f"{abs(v):,.0f}"

    def _fs(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "—"
        try:
            v = float(v)
        except (TypeError, ValueError):
            return "—"
        sign = "+" if v > 0 else ("-" if v < 0 else "")
        return f"{sign}{_sym}{abs(v):,.0f}"

    def _fp(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "—"
        try:
            v = float(v)
        except (TypeError, ValueError):
            return "—"
        sign = "+" if v > 0 else ("-" if v < 0 else "")
        return f"{sign}{abs(v):.1f}%"

    # ── Styles ───────────────────────────────────────────────────────────────
    base = getSampleStyleSheet()

    def _ps(name, **kw):
        return ParagraphStyle(name, parent=base["Normal"], **kw)

    s_body   = _ps("bd",  fontSize=9.5,  textColor=FG2,    leading=14)
    s_bodyb  = _ps("bdb", fontSize=9.5,  textColor=FG2,    leading=14,  fontName="Helvetica-Bold")
    s_cap    = _ps("cp",  fontSize=8.5,  textColor=FG3,    leading=12)
    s_lbl    = _ps("lb",  fontSize=8,    textColor=FG3,    leading=10,  fontName="Helvetica-Bold")
    s_num    = _ps("nm",  fontSize=9.5,  textColor=FG2,    leading=14,  alignment=TA_RIGHT)
    s_numb   = _ps("nmb", fontSize=9.5,  textColor=FG2,    leading=14,  fontName="Helvetica-Bold", alignment=TA_RIGHT)
    s_fav    = _ps("fv",  fontSize=9.5,  textColor=FAV_C,  leading=14,  fontName="Helvetica-Bold")
    s_adv    = _ps("av",  fontSize=9.5,  textColor=ADV_C,  leading=14,  fontName="Helvetica-Bold")
    s_favr   = _ps("fvr", fontSize=9.5,  textColor=FAV_C,  leading=14,  fontName="Helvetica-Bold", alignment=TA_RIGHT)
    s_advr   = _ps("avr", fontSize=9.5,  textColor=ADV_C,  leading=14,  fontName="Helvetica-Bold", alignment=TA_RIGHT)
    s_h2     = _ps("h2",  fontSize=12,   textColor=NAVY,   leading=16,  fontName="Helvetica-Bold",
                   spaceBefore=14, spaceAfter=8)
    s_h3     = _ps("h3",  fontSize=9,    textColor=PRIMARY, leading=12,  fontName="Helvetica-Bold",
                   spaceBefore=12, spaceAfter=6)
    s_navyb  = _ps("nb",  fontSize=9.5,  textColor=NAVY,   leading=14,  fontName="Helvetica-Bold")
    s_white  = _ps("wh",  fontSize=8.5,  textColor=C.white, leading=11)
    s_favsec = _ps("fvh", fontSize=11,   textColor=FAV_C,  leading=15,  fontName="Helvetica-Bold",
                   spaceBefore=12, spaceAfter=6)
    s_advsec = _ps("avh", fontSize=11,   textColor=ADV_C,  leading=15,  fontName="Helvetica-Bold",
                   spaceBefore=12, spaceAfter=6)
    s_capw   = _ps("cpw", fontSize=8,    textColor=C.HexColor("#94A3B8"), leading=11)
    s_lbl_sm = _ps("lbs", fontSize=7.5,  textColor=FG3,    leading=10,  fontName="Helvetica-Bold",
                   alignment=TA_RIGHT)

    # ── Page geometry ────────────────────────────────────────────────────────
    buf = BytesIO()
    A4W, A4H   = A4
    LDW, LDH   = rl_landscape(A4)
    MARG       = 1.5 * cm
    FOOT       = 1.3 * cm   # space at bottom for footer text
    COV_TOP    = 3.3 * cm   # dark header on cover page
    STD_TOP    = 1.4 * cm   # thin header on standard pages

    # frame(x, y_bottom, width, height)
    cover_frame   = Frame(MARG, FOOT, A4W - 2*MARG, A4H - FOOT - COV_TOP - 0.4*cm, id="cover")
    std_frame     = Frame(MARG, FOOT, A4W - 2*MARG, A4H - FOOT - STD_TOP - 0.3*cm, id="std")
    land_frame    = Frame(MARG, FOOT, LDW - 2*MARG, LDH - FOOT - STD_TOP - 0.3*cm, id="land")

    # ── Page decorators ──────────────────────────────────────────────────────
    def _footer(canvas, doc):
        canvas.saveState()
        w, h = doc.pagesize
        canvas.setFillColor(BORDER)
        canvas.rect(MARG, FOOT - 0.15*cm, w - 2*MARG, 0.5, fill=1, stroke=0)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(FG3)
        canvas.drawString(MARG, FOOT - 0.5*cm,
                          f"MonthEndIQ Management Pack  ·  {period_label_str}  ·  {gen_date}")
        canvas.drawRightString(w - MARG, FOOT - 0.5*cm, f"Page {doc.page}")
        canvas.restoreState()

    def _cover_header(canvas, doc):
        _footer(canvas, doc)
        canvas.saveState()
        w, h = doc.pagesize
        # Dark header band
        canvas.setFillColor(NAVY)
        canvas.rect(0, h - COV_TOP, w, COV_TOP, fill=1, stroke=0)
        # Primary accent strip
        canvas.setFillColor(PRIMARY)
        canvas.rect(0, h - COV_TOP - 0.18*cm, w, 0.18*cm, fill=1, stroke=0)
        # MonthEndIQ wordmark
        canvas.setFont("Helvetica-Bold", 16)
        canvas.setFillColor(C.white)
        canvas.drawString(MARG, h - 1.55*cm, "MonthEndIQ")
        # "Management Pack" label
        canvas.setFont("Helvetica-Bold", 13)
        canvas.setFillColor(C.HexColor("#93C5FD"))
        canvas.drawString(MARG + 6.3*cm, h - 1.55*cm, "Management Pack")
        # Sub-line
        canvas.setFont("Helvetica", 9.5)
        canvas.setFillColor(C.HexColor("#94A3B8"))
        analysis_lbl = "Budget vs Actual" if is_bva else "P&L Variance Analysis"
        prepared_by  = f"  ·  Prepared by {firm_name}" if firm_name.strip() else ""
        canvas.drawString(MARG, h - 2.55*cm,
                          f"{period_label_str}  ·  {analysis_lbl}  ·  {gen_date}{prepared_by}")
        canvas.restoreState()

    def _std_header(canvas, doc):
        _footer(canvas, doc)
        canvas.saveState()
        w, h = doc.pagesize
        canvas.setFillColor(NAVY)
        canvas.rect(0, h - STD_TOP, w, STD_TOP, fill=1, stroke=0)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.setFillColor(C.white)
        canvas.drawString(MARG, h - 0.9*cm, "MonthEndIQ Management Pack")
        canvas.setFont("Helvetica", 8.5)
        canvas.setFillColor(C.HexColor("#94A3B8"))
        canvas.drawRightString(w - MARG, h - 0.9*cm, period_label_str)
        canvas.restoreState()

    templates = [
        PageTemplate(id="cover",    frames=[cover_frame], onPage=_cover_header, pagesize=A4),
        PageTemplate(id="standard", frames=[std_frame],   onPage=_std_header,   pagesize=A4),
        PageTemplate(id="landscape",frames=[land_frame],  onPage=_std_header,   pagesize=rl_landscape(A4)),
    ]
    doc = BaseDocTemplate(buf, pageTemplates=templates, pagesize=A4)

    story = []
    Sp = lambda n: Spacer(1, n * cm)

    # ────────────────────────────────────────────────────────────────────────
    # PAGE 1 — Cover: KPIs · Executive Summary · Commentary · Actions
    # ────────────────────────────────────────────────────────────────────────
    AW = A4W - 2 * MARG   # available content width

    story.append(Sp(0.4))
    story.append(Paragraph("KPI SUMMARY", s_h3))

    # KPI table
    kpi_header = [
        Paragraph("Metric",      s_lbl),
        Paragraph(actual_word,   s_lbl),
        Paragraph(prior_word,    s_lbl),
        Paragraph("Variance",    s_lbl),
        Paragraph("Change %",    s_lbl),
    ]
    kpi_data = [kpi_header]
    for k in kpis:
        if k.get("pct_only"):
            pct = k.get("pct")
            p_s = s_favr if k.get("is_fav") else s_advr
            kpi_data.append([
                Paragraph(k.get("label", ""), s_bodyb),
                Paragraph("—", s_num), Paragraph("—", s_num),
                Paragraph("—", s_num), Paragraph(_fp(pct), p_s),
            ])
        else:
            fav = k.get("is_fav", True)
            p_s = s_favr if fav else s_advr
            kpi_data.append([
                Paragraph(k.get("label", ""), s_bodyb),
                Paragraph(_f(k.get("value")),    s_num),
                Paragraph(_f(k.get("prior")),    s_num),
                Paragraph(_fs(k.get("variance")), p_s),
                Paragraph(_fp(k.get("pct")),     p_s),
            ])

    cw = [AW * p for p in (0.30, 0.175, 0.175, 0.175, 0.175)]
    kpi_tbl = Table(kpi_data, colWidths=cw)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), SOFT),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 8),
        ("TEXTCOLOR",     (0, 0), (-1, 0), FG3),
        ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN",         (0, 0), (0, -1), "LEFT"),
        ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C.white, SOFT]),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("FONTSIZE",      (0, 1), (-1, -1), 9.5),
    ]))
    story.append(kpi_tbl)
    story.append(Sp(0.5))

    # Executive Summary
    if report["executive_summary"]:
        story.append(Paragraph("EXECUTIVE SUMMARY", s_h3))
        for line in report["executive_summary"]:
            story.append(Paragraph(line, s_body))
        story.append(Sp(0.35))

    # Board Commentary
    if report["commentary"]:
        story.append(Paragraph("BOARD COMMENTARY", s_h3))
        for line in report["commentary"]:
            story.append(Paragraph(f"• {line}", s_body))
        story.append(Sp(0.35))

    # Recommended Actions
    if report["actions"]:
        story.append(Paragraph("RECOMMENDED ACTIONS", s_h3))
        for line in report["actions"]:
            if line.startswith("[HIGH]"):
                pill_txt, pill_c, pill_bg, border_c = "HIGH",   ADV_C,  ADV_BG,   ADV_C
                body_txt = line[7:].strip()
            elif line.startswith("[MEDIUM]"):
                pill_txt, pill_c, pill_bg, border_c = "MEDIUM", AMBER,  AMBER_BG, AMBER
                body_txt = line[9:].strip()
            else:
                pill_txt, pill_c, pill_bg, border_c = "LOW",    FAV_C,  FAV_BG,   FAV_C
                body_txt = line[6:].strip()
            s_pill = _ps(f"pill_{pill_txt}", fontSize=7.5, fontName="Helvetica-Bold",
                         textColor=pill_c, leading=10, alignment=TA_LEFT)
            row = [[Paragraph(pill_txt, s_pill), Paragraph(body_txt, s_body)]]
            at = Table(row, colWidths=[2.0*cm, AW - 2.0*cm])
            at.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, -1), SOFT),
                ("BACKGROUND",   (0, 0), (0,  0),  pill_bg),
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",   (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 7),
                ("LEFTPADDING",  (0, 0), (0,  0),  6),
                ("RIGHTPADDING", (0, 0), (0,  0),  6),
                ("LEFTPADDING",  (1, 0), (1, -1),  10),
                ("RIGHTPADDING", (1, 0), (1, -1),  8),
                ("LINEBEFORE",   (0, 0), (0, -1), 3, border_c),
                ("BOX",          (0, 0), (-1, -1), 0.3, BORDER),
            ]))
            story.append(at)
            story.append(Sp(0.15))

    # ────────────────────────────────────────────────────────────────────────
    # PAGE 2 — Key Findings: Fav · Adv · Profit Drivers
    # ────────────────────────────────────────────────────────────────────────
    story.append(NextPageTemplate("standard"))
    story.append(PageBreak())
    story.append(Paragraph("Key Findings", _ps("kfh", fontSize=15, textColor=NAVY,
                                               fontName="Helvetica-Bold", leading=19, spaceAfter=12)))

    def _var_table(rows_data, bg_hdr):
        hdr = [Paragraph(t, s_lbl) for t in ["Account", "Category", actual_word, prior_word, "Variance", "Var %"]]
        tbl_data = [hdr]
        for m in rows_data:
            fav = m.get("is_fav", True)
            rs = s_favr if fav else s_advr
            tbl_data.append([
                Paragraph(m["account"][:48], s_body),
                Paragraph(m["category"],     s_cap),
                Paragraph(_f(m.get("value")),       s_num),
                Paragraph(_f(m.get("prior_value")), s_num),
                Paragraph(_fs(m.get("variance")),   rs),
                Paragraph(_fp(m.get("variance_pct")), rs),
            ])
        t = Table(tbl_data, colWidths=[AW*0.31, AW*0.18, AW*0.13, AW*0.13, AW*0.13, AW*0.12])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), bg_hdr),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 8),
            ("ALIGN",         (2, 0), (-1, -1), "RIGHT"),
            ("ALIGN",         (0, 0), (1, -1), "LEFT"),
            ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C.white, SOFT]),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("FONTSIZE",      (0, 1), (-1, -1), 9),
        ]))
        return t

    # Favourable
    story.append(Paragraph("▲  Favourable Variances", s_favsec))
    if report["key_findings_fav"]:
        story.append(_var_table(report["key_findings_fav"], FAV_BG))
    else:
        story.append(Paragraph("No favourable variances in this period.", s_body))
    story.append(Sp(0.4))

    # Adverse
    story.append(Paragraph("▼  Adverse Variances", s_advsec))
    if report["key_findings_adv"]:
        story.append(_var_table(report["key_findings_adv"], ADV_BG))
    else:
        story.append(Paragraph("No adverse variances in this period.", s_body))
    story.append(Sp(0.4))

    # Profit drivers
    if waterfall:
        story.append(Paragraph("Profit Drivers", _ps("pdh", fontSize=11, textColor=NAVY,
                                                      fontName="Helvetica-Bold", leading=15,
                                                      spaceBefore=10, spaceAfter=6)))
        prior_lbl_wf   = "Budget Profit"  if is_bva else "Prior Period Profit"
        current_lbl_wf = "Actual Profit"  if is_bva else "Current Period Profit"

        wf_data = [[Paragraph(prior_lbl_wf, s_bodyb), Paragraph(_f(waterfall["prior_profit"]), s_numb)]]
        for b in waterfall.get("bars", []):
            rs = s_favr if b.get("fav") else s_advr
            wf_data.append([Paragraph(b["label"], s_body), Paragraph(_fs(b["impact"]), rs)])
        wf_data.append([Paragraph(current_lbl_wf, s_bodyb), Paragraph(_f(waterfall["current_profit"]), s_numb)])
        net = waterfall.get("net_change", 0) or 0
        net_rs = s_favr if net >= 0 else s_advr
        wf_data.append([Paragraph("Net Change", s_navyb), Paragraph(_fs(net), net_rs)])

        wf_t = Table(wf_data, colWidths=[AW * 0.65, AW * 0.35])
        wf_style = [
            ("ALIGN",         (1, 0), (1, -1), "RIGHT"),
            ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("BACKGROUND",    (0, 0), (-1, 0), SOFT),
            ("LINEABOVE",     (0, -1), (-1, -1), 1, NAVY),
            ("TOPPADDING",    (0, -1), (-1, -1), 8),
        ]
        for i, b in enumerate(waterfall.get("bars", []), start=1):
            wf_style.append(("BACKGROUND", (0, i), (-1, i), FAV_BG if b.get("fav") else ADV_BG))
        n = len(wf_data)
        wf_style.append(("BACKGROUND", (0, n-2), (-1, n-2), SOFT))
        wf_t.setStyle(TableStyle(wf_style))
        story.append(wf_t)

    # ────────────────────────────────────────────────────────────────────────
    # PAGE 3 — Variance Analysis (landscape)
    # ────────────────────────────────────────────────────────────────────────
    story.append(NextPageTemplate("landscape"))
    story.append(PageBreak())
    story.append(Paragraph(
        f"Variance Analysis — {period_label_str}",
        _ps("vah", fontSize=13, textColor=NAVY, fontName="Helvetica-Bold", leading=17, spaceAfter=10),
    ))

    LW = LDW - 2 * MARG
    land_cw = [LW*p for p in (0.26, 0.14, 0.12, 0.12, 0.12, 0.09, 0.10, 0.05)]
    land_hdr = [
        Paragraph("Account",        s_white),
        Paragraph("Category",       s_white),
        Paragraph(actual_word,      s_white),
        Paragraph(prior_word,       s_white),
        Paragraph("Variance",       s_white),
        Paragraph("Var %",          s_white),
        Paragraph("Impact",         s_white),
        Paragraph("",               s_white),
    ]
    land_data = [land_hdr]
    for m in movements[:28]:
        fav = m.get("is_fav", True)
        rs  = s_favr if fav else s_advr
        impact_lbl = _ps(f"imp_{id(m)}", fontSize=8, textColor=FAV_C if fav else ADV_C,
                         fontName="Helvetica-Bold", leading=11)
        land_data.append([
            Paragraph(m["account"][:46],              s_body),
            Paragraph(m["category"],                  s_cap),
            Paragraph(_f(m.get("value")),             s_num),
            Paragraph(_f(m.get("prior_value")),       s_num),
            Paragraph(_fs(m.get("variance")),         rs),
            Paragraph(_fp(m.get("variance_pct")),     rs),
            Paragraph("Fav" if fav else "Adv",        impact_lbl),
            Paragraph("",                             s_body),
        ])

    land_tbl = Table(land_data, colWidths=land_cw)
    land_style = [
        ("BACKGROUND",   (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",    (0, 0), (-1, 0), C.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 8),
        ("ALIGN",        (2, 0), (-1, -1), "RIGHT"),
        ("ALIGN",        (0, 0), (1, -1), "LEFT"),
        ("GRID",         (0, 0), (-1, -1), 0.3, BORDER),
        ("FONTSIZE",     (0, 1), (-1, -1), 8.5),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]
    for i, m in enumerate(movements[:28], start=1):
        bg = FAV_BG if m.get("is_fav") else (ADV_BG if m.get("variance", 0) != 0 else C.white)
        land_style.append(("BACKGROUND", (0, i), (-1, i), bg))

    land_tbl.setStyle(TableStyle(land_style))
    story.append(land_tbl)

    # ────────────────────────────────────────────────────────────────────────
    # PAGE 4 — Insights Summary (only when insights data is provided)
    # ────────────────────────────────────────────────────────────────────────
    if insights:
        story.append(NextPageTemplate("standard"))
        story.append(PageBreak())
        story.append(Paragraph(
            "Insights Summary",
            _ps("ish", fontSize=15, textColor=NAVY, fontName="Helvetica-Bold", leading=19, spaceAfter=8),
        ))

        # Helper: two-column label/value row builder
        def _ins_row(label, value, val_style=None):
            return [
                Paragraph(label, s_body),
                Paragraph(value, val_style or s_numb),
            ]

        def _ins_section(title):
            story.append(Paragraph(title, s_h3))

        # ── Operating Margins ────────────────────────────────────────────
        _ins_section("OPERATING MARGINS")
        margins = insights.get("margins", {})
        op_pct  = margins.get("op_pct")
        pay_pct = margins.get("payroll_pct")

        # Direction from trend
        trend = margins.get("trend", [])
        op_dir = "—"
        if len(trend) >= 3:
            recent = [t["op_pct"] for t in trend[-3:] if t.get("op_pct") is not None]
            if len(recent) >= 2:
                chg = recent[-1] - recent[0]
                op_dir = "Improving" if chg > 0.5 else ("Worsening" if chg < -0.5 else "Stable")

        margin_data = [
            [Paragraph("Metric", s_lbl), Paragraph("Value", s_lbl), Paragraph("Trend", s_lbl)],
            [
                Paragraph("Operating Margin", s_body),
                Paragraph(f"{op_pct:.1f}%" if op_pct is not None else "—", s_numb),
                Paragraph(op_dir, s_fav if op_dir == "Improving" else (s_adv if op_dir == "Worsening" else s_body)),
            ],
            [
                Paragraph("Payroll % of Revenue", s_body),
                Paragraph(f"{pay_pct:.1f}%" if pay_pct is not None else "—", s_numb),
                Paragraph("", s_body),
            ],
        ]
        m_tbl = Table(margin_data, colWidths=[AW * 0.45, AW * 0.25, AW * 0.30])
        m_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), SOFT),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 8),
            ("TEXTCOLOR",     (0, 0), (-1, 0), FG3),
            ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN",         (0, 0), (0, -1), "LEFT"),
            ("ALIGN",         (2, 0), (2, -1), "LEFT"),
            ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C.white, SOFT]),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ]))
        story.append(m_tbl)
        story.append(Sp(0.4))

        # ── Projections: R12 & Run-Rate ──────────────────────────────────
        _ins_section("PROJECTIONS")
        r12      = insights.get("r12", {})
        run_rate = insights.get("run_rate", {})

        proj_hdr = [Paragraph(h, s_lbl) for h in ["Metric", "R12 (Rolling 12 Mo)", "Run-Rate (Annualised)"]]
        proj_data = [proj_hdr]
        for key, label in (("revenue", "Revenue"), ("costs", "Total Costs"), ("profit", "Profit")):
            r12_val  = r12.get(key)  if r12.get("available") else None
            rr_val   = run_rate.get(key)
            proj_data.append([
                Paragraph(label, s_body),
                Paragraph(_f(r12_val),  s_numb),
                Paragraph(_f(rr_val),   s_numb),
            ])
        proj_tbl = Table(proj_data, colWidths=[AW * 0.35, AW * 0.325, AW * 0.325])
        proj_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), SOFT),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 8),
            ("TEXTCOLOR",     (0, 0), (-1, 0), FG3),
            ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN",         (0, 0), (0, -1), "LEFT"),
            ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C.white, SOFT]),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ]))
        story.append(proj_tbl)
        story.append(Sp(0.4))

        # ── Cost Pareto (Top 5) ──────────────────────────────────────────
        _ins_section("COST CONCENTRATION (TOP 5 LINES)")
        pareto = insights.get("pareto", {})
        top5   = (pareto.get("lines") or [])[:5]

        if top5:
            pareto_hdr = [Paragraph(h, s_lbl) for h in ["Account", "Category", "Value", "% of Total Costs", "Cumulative %"]]
            pareto_data = [pareto_hdr]
            for pl in top5:
                pareto_data.append([
                    Paragraph(str(pl.get("account", ""))[:44], s_body),
                    Paragraph(str(pl.get("category", "")),     s_cap),
                    Paragraph(_f(pl.get("value")),             s_num),
                    Paragraph(f"{pl.get('pct_of_total', 0):.1f}%", s_num),
                    Paragraph(f"{pl.get('cum_pct', 0):.1f}%",      s_num),
                ])
            par_tbl = Table(pareto_data, colWidths=[AW*0.32, AW*0.18, AW*0.16, AW*0.17, AW*0.17])
            par_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), SOFT),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, 0), 8),
                ("TEXTCOLOR",     (0, 0), (-1, 0), FG3),
                ("ALIGN",         (2, 0), (-1, -1), "RIGHT"),
                ("ALIGN",         (0, 0), (1, -1), "LEFT"),
                ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C.white, SOFT]),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                ("FONTSIZE",      (0, 1), (-1, -1), 9),
            ]))
            story.append(par_tbl)
        else:
            story.append(Paragraph("No cost data available.", s_body))
        story.append(Sp(0.4))

        # ── Momentum ─────────────────────────────────────────────────────
        _ins_section("MOMENTUM (3-MONTH TREND)")
        momentum = insights.get("momentum", {})
        if momentum.get("available"):
            def _dir_style(d):
                return s_fav if d == "improving" else (s_adv if d == "worsening" else s_body)

            mom_data = [
                [Paragraph(h, s_lbl) for h in ["Metric", "Direction"]],
                [Paragraph("Revenue",      s_body), Paragraph(str(momentum.get("revenue_dir", "—")).capitalize(), _dir_style(momentum.get("revenue_dir")))],
                [Paragraph("Total Costs",  s_body), Paragraph(str(momentum.get("cost_dir",    "—")).capitalize(), _dir_style(momentum.get("cost_dir")))],
                [Paragraph("Profit",       s_body), Paragraph(str(momentum.get("profit_dir",  "—")).capitalize(), _dir_style(momentum.get("profit_dir")))],
                [Paragraph("Overall",      s_bodyb), Paragraph(str(momentum.get("overall",    "—")).capitalize(), _dir_style(momentum.get("overall")))],
            ]
            mom_tbl = Table(mom_data, colWidths=[AW * 0.55, AW * 0.45])
            mom_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), SOFT),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, 0), 8),
                ("TEXTCOLOR",     (0, 0), (-1, 0), FG3),
                ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
                ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C.white, SOFT]),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                ("LINEABOVE",     (0, -1), (-1, -1), 0.5, NAVY),
            ]))
            story.append(mom_tbl)
        else:
            story.append(Paragraph("Insufficient data for momentum analysis (requires 6+ periods).", s_body))
        story.append(Sp(0.4))

        # ── Same-Period Prior Year ────────────────────────────────────────
        sppy = insights.get("sppy", {})
        if sppy.get("available"):
            _ins_section(f"SAME-PERIOD PRIOR YEAR ({sppy.get('period_label', '')})")
            sppy_data = [
                [Paragraph(h, s_lbl) for h in ["Metric", "Current", "Prior Year", "Change", "Change %"]],
                [
                    Paragraph("Revenue", s_body),
                    Paragraph(_f((margins.get("trend") or [{}])[-1].get("revenue")), s_num),
                    Paragraph(_f(sppy.get("revenue")),    s_num),
                    Paragraph(_fs(sppy.get("rev_delta")), s_favr if (sppy.get("rev_delta") or 0) >= 0 else s_advr),
                    Paragraph(_fp(sppy.get("rev_pct")),   s_favr if (sppy.get("rev_pct") or 0) >= 0 else s_advr),
                ],
                [
                    Paragraph("Profit", s_body),
                    Paragraph(_f((margins.get("trend") or [{}])[-1].get("profit")), s_num),
                    Paragraph(_f(sppy.get("profit")),     s_num),
                    Paragraph(_fs(sppy.get("prof_delta")), s_favr if (sppy.get("prof_delta") or 0) >= 0 else s_advr),
                    Paragraph(_fp(sppy.get("prof_pct")),   s_favr if (sppy.get("prof_pct") or 0) >= 0 else s_advr),
                ],
            ]
            sppy_tbl = Table(sppy_data, colWidths=[AW*0.25, AW*0.175, AW*0.175, AW*0.175, AW*0.175])
            sppy_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), SOFT),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, 0), 8),
                ("TEXTCOLOR",     (0, 0), (-1, 0), FG3),
                ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
                ("ALIGN",         (0, 0), (0, -1), "LEFT"),
                ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C.white, SOFT]),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ]))
            story.append(sppy_tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


def make_xlsx(period_label_str: str, movements: list, commentary: list, kpis: list,
              analysis_type: str = "month_on_month",
              insights: dict | None = None,
              nhs_data: dict | None = None) -> bytes:
    """Generate a formatted Excel workbook with sheets: Variance, KPIs, Commentary, Insights.

    When nhs_data is provided (sector == 'nhs_gp'), also writes:
      • ICB Variance Report  — commissioner-ready structured variance table
      • Workforce            — staff cost breakdown by role type
      • Partner Accounts     — net profit, drawings, surplus, tax reserve
      • AISMA Mapping        — P&L mapped to standard NHS/AISMA headings
    """
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
    )
    from openpyxl.utils import get_column_letter

    is_bva    = analysis_type == "budget_vs_actual"
    prior_word = "Budget" if is_bva else "Prior"

    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────────────────────────────
    C_HEADER    = "1A2B4C"   # navy header
    C_FAV_BG    = "ECFDF5"   # green tint
    C_ADV_BG    = "FEF2F2"   # red tint
    C_FAV_FG    = "065F46"
    C_ADV_FG    = "991B1B"
    C_SUBHEAD   = "F1F5F9"   # light grey for sub-header rows
    C_WHITE     = "FFFFFF"

    def hdr_font(bold=True): return Font(name="Calibri", bold=bold, color="FFFFFF", size=10)
    def body_font(bold=False, color="1A1A1A"): return Font(name="Calibri", bold=bold, color=color, size=10)
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def thin_border():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)
    def gbp_fmt(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return ""
        return v
    def pct_fmt(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return ""
        return round(v, 1)

    GBP_FMT   = '#,##0;[Red]-#,##0'
    GBP_SFMT  = '+#,##0;-#,##0;"-"'
    PCT_FMT   = '+0.0%;-0.0%;"-"'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Variance Analysis
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Variance Analysis"

    # Title row
    ws1.merge_cells("A1:G1")
    t = ws1["A1"]
    t.value = f"Variance Analysis — {period_label_str}"
    t.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 24

    # Column headers
    headers = ["Account", "Category", "Current" if not is_bva else "Actual",
               prior_word, "Variance", "Var %", "Impact"]
    for col, h in enumerate(headers, start=1):
        c = ws1.cell(row=2, column=col, value=h)
        c.font      = hdr_font()
        c.fill      = fill(C_HEADER)
        c.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
        c.border    = thin_border()
    ws1.row_dimensions[2].height = 18

    # Data rows
    for row_i, m in enumerate(movements, start=3):
        is_fav = m.get("is_fav", False)
        var    = m.get("variance", 0) or 0
        is_zero = var == 0

        row_bg = C_FAV_BG if is_fav else (C_ADV_BG if not is_zero else C_WHITE)
        impact_txt = "Favourable" if is_fav else ("Adverse" if not is_zero else "—")
        impact_col = C_FAV_FG if is_fav else (C_ADV_FG if not is_zero else "6B7280")

        vals = [
            m.get("account", ""),
            m.get("category", ""),
            gbp_fmt(m.get("value")),
            gbp_fmt(m.get("prior_value")),
            gbp_fmt(var),
            pct_fmt(m.get("variance_pct")),
            impact_txt,
        ]
        fmts = [None, None, GBP_FMT, GBP_FMT, GBP_SFMT, PCT_FMT, None]
        for col, (v, f) in enumerate(zip(vals, fmts), start=1):
            c = ws1.cell(row=row_i, column=col, value=v)
            c.fill      = fill(row_bg)
            c.border    = thin_border()
            c.alignment = Alignment(
                horizontal="left" if col <= 2 else "right",
                vertical="center",
            )
            if col == 7:  # Impact badge column
                c.font = body_font(bold=True, color=impact_col)
            elif col in (5, 6):
                fav_c = C_FAV_FG if is_fav else (C_ADV_FG if not is_zero else "6B7280")
                c.font = body_font(bold=True, color=fav_c)
                if f: c.number_format = f
            else:
                c.font = body_font()
                if f: c.number_format = f

    # Column widths
    col_widths = [34, 22, 14, 14, 14, 10, 13]
    for i, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.freeze_panes = "A3"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — KPI Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("KPI Summary")
    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value = f"KPI Summary — {period_label_str}"
    t2.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 24

    kpi_headers = ["Metric", "Value", prior_word, "Variance", "Var %"]
    for col, h in enumerate(kpi_headers, start=1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font      = hdr_font()
        c.fill      = fill(C_HEADER)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        c.border    = thin_border()
    ws2.row_dimensions[2].height = 18

    for row_i, k in enumerate([k for k in kpis if not k.get("pct_only")], start=3):
        is_fav  = k.get("is_fav", False)
        var     = k.get("variance", 0) or 0
        row_bg  = C_FAV_BG if is_fav else (C_ADV_BG if var != 0 else C_WHITE)
        kpi_vals = [k.get("label",""), gbp_fmt(k.get("value")), gbp_fmt(k.get("prior")),
                    gbp_fmt(var), pct_fmt(k.get("pct"))]
        kpi_fmts = [None, GBP_FMT, GBP_FMT, GBP_SFMT, PCT_FMT]
        for col, (v, f) in enumerate(zip(kpi_vals, kpi_fmts), start=1):
            c = ws2.cell(row=row_i, column=col, value=v)
            c.fill   = fill(row_bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
            c.font   = body_font(bold=(col in (4, 5)), color=(C_FAV_FG if (is_fav and col in (4,5)) else (C_ADV_FG if (not is_fav and var != 0 and col in (4,5)) else "1A1A1A")))
            if f: c.number_format = f

    for i, w in enumerate([28, 16, 16, 16, 12], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Commentary
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Commentary")
    ws3.merge_cells("A1:B1")
    t3 = ws3["A1"]
    t3.value = f"Management Commentary — {period_label_str}"
    t3.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 24

    import html as html_lib
    for row_i, c_item in enumerate(commentary, start=2):
        raw = html_lib.unescape(
            c_item.get("html", "").replace("<b>", "").replace("</b>", "")
            .replace("<br>", "\n").replace("<br/>", "\n")
        )
        import re as _re
        raw = _re.sub(r"<[^>]+>", "", raw).strip()

        icon_col = "065F46" if c_item.get("fav") else "991B1B"
        bullet = ws3.cell(row=row_i, column=1, value="●")
        bullet.font = Font(name="Calibri", size=10, color=icon_col, bold=True)
        bullet.alignment = Alignment(horizontal="center", vertical="top")
        bullet.border = thin_border()

        txt = ws3.cell(row=row_i, column=2, value=raw)
        txt.font = body_font()
        txt.alignment = Alignment(wrap_text=True, vertical="top")
        txt.border = thin_border()
        ws3.row_dimensions[row_i].height = max(15, len(raw) // 6)

    ws3.column_dimensions["A"].width = 4
    ws3.column_dimensions["B"].width = 90

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Insights (only when insights data is provided)
    # ══════════════════════════════════════════════════════════════════════════
    if insights:
        ws4 = wb.create_sheet("Insights")

        # Title
        ws4.merge_cells("A1:D1")
        t4 = ws4["A1"]
        t4.value = f"Insights Summary — {period_label_str}"
        t4.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
        t4.alignment = Alignment(horizontal="left", vertical="center")
        ws4.row_dimensions[1].height = 24

        cur_row = 2

        def _ws4_section(title):
            nonlocal cur_row
            ws4.merge_cells(f"A{cur_row}:D{cur_row}")
            c = ws4.cell(row=cur_row, column=1, value=title)
            c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill      = fill(C_HEADER)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws4.row_dimensions[cur_row].height = 16
            cur_row += 1

        def _ws4_row(label, *values, bold_label=False):
            nonlocal cur_row
            c = ws4.cell(row=cur_row, column=1, value=label)
            c.font      = body_font(bold=bold_label)
            c.fill      = fill(C_SUBHEAD if bold_label else C_WHITE)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = thin_border()
            for col_i, val in enumerate(values, start=2):
                cv = ws4.cell(row=cur_row, column=col_i, value=val)
                cv.font      = body_font(bold=bold_label)
                cv.fill      = fill(C_SUBHEAD if bold_label else C_WHITE)
                cv.border    = thin_border()
                cv.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left",
                                         vertical="center")
                if isinstance(val, (int, float)):
                    cv.number_format = GBP_FMT
            cur_row += 1

        def _ws4_pct_row(label, pct_value, note=""):
            nonlocal cur_row
            c = ws4.cell(row=cur_row, column=1, value=label)
            c.font = body_font(); c.fill = fill(C_WHITE); c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center")
            pv = ws4.cell(row=cur_row, column=2, value=round(pct_value, 1) if pct_value is not None else None)
            pv.font = body_font(); pv.fill = fill(C_WHITE); pv.border = thin_border()
            pv.alignment = Alignment(horizontal="right", vertical="center")
            if pv.value is not None:
                pv.number_format = '0.0"%"'
            if note:
                nv = ws4.cell(row=cur_row, column=3, value=note)
                nv.font = body_font(color="667085"); nv.fill = fill(C_WHITE); nv.border = thin_border()
                nv.alignment = Alignment(horizontal="left", vertical="center")
            cur_row += 1

        def _ws4_text_row(label, value, fav=None):
            nonlocal cur_row
            fg = C_FAV_FG if fav is True else (C_ADV_FG if fav is False else "1A1A1A")
            c = ws4.cell(row=cur_row, column=1, value=label)
            c.font = body_font(); c.fill = fill(C_WHITE); c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center")
            v = ws4.cell(row=cur_row, column=2, value=value)
            v.font = body_font(bold=(fav is not None), color=fg)
            v.fill = fill(C_FAV_BG if fav is True else (C_ADV_BG if fav is False else C_WHITE))
            v.border = thin_border()
            v.alignment = Alignment(horizontal="left", vertical="center")
            cur_row += 1

        # ── Section 1: Operating Margins ──────────────────────────────
        _ws4_section("Operating Margins")

        margins  = insights.get("margins", {})
        op_pct   = margins.get("op_pct")
        pay_pct  = margins.get("payroll_pct")

        # Compute operating margin direction from trend
        trend    = margins.get("trend", [])
        op_dir   = "Stable"
        if len(trend) >= 3:
            recent = [t["op_pct"] for t in trend[-3:] if t.get("op_pct") is not None]
            if len(recent) >= 2:
                chg    = recent[-1] - recent[0]
                op_dir = "Improving" if chg > 0.5 else ("Worsening" if chg < -0.5 else "Stable")

        _ws4_pct_row("Operating Margin %", op_pct, op_dir)
        _ws4_pct_row("Payroll % of Revenue", pay_pct)
        cur_row += 1  # blank spacer

        # ── Section 2: Projections ────────────────────────────────────
        _ws4_section("Projections")

        r12      = insights.get("r12", {})
        run_rate = insights.get("run_rate", {})

        # Column headers for projections sub-table
        for col_i, hdr in enumerate(["Metric", "R12 (Rolling 12 Mo)", "Run-Rate (Annualised)"], start=1):
            hc = ws4.cell(row=cur_row, column=col_i, value=hdr)
            hc.font = Font(name="Calibri", bold=True, size=9, color="667085")
            hc.fill = fill(C_SUBHEAD)
            hc.border = thin_border()
            hc.alignment = Alignment(horizontal="right" if col_i > 1 else "left", vertical="center")
        cur_row += 1

        for key, label in (("revenue", "Revenue"), ("costs", "Total Costs"), ("profit", "Profit")):
            r12_val = r12.get(key) if r12.get("available") else None
            rr_val  = run_rate.get(key)
            _ws4_row(label, r12_val, rr_val)
        cur_row += 1  # blank spacer

        # ── Section 3: Cost Pareto (Top 10) ──────────────────────────
        _ws4_section("Cost Concentration — Top 10 Lines")

        pareto  = insights.get("pareto", {})
        top10   = (pareto.get("lines") or [])[:10]

        if top10:
            for col_i, hdr in enumerate(["Account", "Category", "Value", "% of Total", "Cumulative %"], start=1):
                hc = ws4.cell(row=cur_row, column=col_i, value=hdr)
                hc.font = Font(name="Calibri", bold=True, size=9, color="667085")
                hc.fill = fill(C_SUBHEAD)
                hc.border = thin_border()
                hc.alignment = Alignment(horizontal="right" if col_i > 2 else "left", vertical="center")
            cur_row += 1
            for i, pl in enumerate(top10):
                row_bg = C_WHITE if i % 2 == 0 else C_SUBHEAD
                vals = [
                    (str(pl.get("account", "")), "left"),
                    (str(pl.get("category", "")), "left"),
                    (pl.get("value"), "right"),
                    (pl.get("pct_of_total"), "right"),
                    (pl.get("cum_pct"), "right"),
                ]
                for col_i, (val, align) in enumerate(vals, start=1):
                    pc = ws4.cell(row=cur_row, column=col_i, value=val)
                    pc.font = body_font()
                    pc.fill = fill(row_bg)
                    pc.border = thin_border()
                    pc.alignment = Alignment(horizontal=align, vertical="center")
                    if col_i == 3 and isinstance(val, (int, float)):
                        pc.number_format = GBP_FMT
                    elif col_i in (4, 5) and isinstance(val, (int, float)):
                        pc.number_format = '0.0"%"'
                cur_row += 1
        else:
            ws4.cell(row=cur_row, column=1, value="No cost data available.").font = body_font()
            cur_row += 1
        cur_row += 1  # blank spacer

        # ── Section 4: Momentum ───────────────────────────────────────
        _ws4_section("Momentum (3-Month Trend)")

        momentum = insights.get("momentum", {})
        if momentum.get("available"):
            def _dir_fav(d):
                return True if d == "improving" else (False if d == "worsening" else None)

            _ws4_text_row("Revenue",     str(momentum.get("revenue_dir", "—")).capitalize(), _dir_fav(momentum.get("revenue_dir")))
            _ws4_text_row("Total Costs", str(momentum.get("cost_dir",    "—")).capitalize(), _dir_fav(momentum.get("cost_dir")))
            _ws4_text_row("Profit",      str(momentum.get("profit_dir",  "—")).capitalize(), _dir_fav(momentum.get("profit_dir")))
            _ws4_text_row("Overall",     str(momentum.get("overall",     "—")).capitalize(), _dir_fav(momentum.get("overall")))
        else:
            ws4.cell(row=cur_row, column=1, value="Insufficient data (requires 6+ periods).").font = body_font()
            cur_row += 1
        cur_row += 1  # blank spacer

        # ── Section 5: Same-Period Prior Year ─────────────────────────
        sppy = insights.get("sppy", {})
        if sppy.get("available"):
            _ws4_section(f"Same-Period Prior Year ({sppy.get('period_label', 'PY')})")

            for col_i, hdr in enumerate(["Metric", "Prior Year", "Change", "Change %"], start=1):
                hc = ws4.cell(row=cur_row, column=col_i, value=hdr)
                hc.font = Font(name="Calibri", bold=True, size=9, color="667085")
                hc.fill = fill(C_SUBHEAD)
                hc.border = thin_border()
                hc.alignment = Alignment(horizontal="right" if col_i > 1 else "left", vertical="center")
            cur_row += 1

            for label, py_key, delta_key, pct_key in (
                ("Revenue", "revenue", "rev_delta",  "rev_pct"),
                ("Profit",  "profit",  "prof_delta", "prof_pct"),
            ):
                py_val    = sppy.get(py_key)
                delta_val = sppy.get(delta_key)
                pct_val   = sppy.get(pct_key)
                is_fav    = (delta_val or 0) >= 0
                row_bg    = C_FAV_BG if is_fav else C_ADV_BG
                fg_col    = C_FAV_FG if is_fav else C_ADV_FG

                cv_data = [
                    (label,     "1A1A1A", C_WHITE, "left"),
                    (py_val,    "1A1A1A", C_WHITE, "right"),
                    (delta_val, fg_col,   row_bg,  "right"),
                    (pct_val,   fg_col,   row_bg,  "right"),
                ]
                for col_i, (val, fc, bg, align) in enumerate(cv_data, start=1):
                    sc = ws4.cell(row=cur_row, column=col_i, value=val)
                    sc.font = Font(name="Calibri", size=10, color=fc,
                                   bold=(col_i in (3, 4)))
                    sc.fill = fill(bg)
                    sc.border = thin_border()
                    sc.alignment = Alignment(horizontal=align, vertical="center")
                    if col_i == 2 and isinstance(val, (int, float)):
                        sc.number_format = GBP_FMT
                    elif col_i == 3 and isinstance(val, (int, float)):
                        sc.number_format = GBP_SFMT
                    elif col_i == 4 and isinstance(val, (int, float)):
                        sc.number_format = '+0.0%;-0.0%;"-"'
                cur_row += 1

        # Column widths for insights sheet
        ws4.column_dimensions["A"].width = 38
        ws4.column_dimensions["B"].width = 20
        ws4.column_dimensions["C"].width = 20
        ws4.column_dimensions["D"].width = 16
        ws4.column_dimensions["E"].width = 16

    # ══════════════════════════════════════════════════════════════════════════
    # NHS SHEETS (only when nhs_data supplied)
    # ══════════════════════════════════════════════════════════════════════════
    if nhs_data:
        re_html = re.compile(r"<[^>]+>")

        # ── ICB Variance Report ───────────────────────────────────────────────
        ws_icb = wb.create_sheet("ICB Variance Report")
        ws_icb.merge_cells("A1:G1")
        hdr = ws_icb["A1"]
        hdr.value = f"ICB Variance Report — {period_label_str}"
        hdr.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
        hdr.alignment = Alignment(horizontal="left", vertical="center")
        ws_icb.row_dimensions[1].height = 22

        icb_headers = ["Account", "Category", "Current (£)", "Prior (£)", "Variance (£)", "Var %", "Commentary"]
        for ci, h in enumerate(icb_headers, 1):
            c = ws_icb.cell(row=2, column=ci, value=h)
            c.font  = hdr_font()
            c.fill  = fill(C_HEADER)
            c.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center")
        ws_icb.row_dimensions[2].height = 18

        icb_row = 3
        # Group movements by category
        from itertools import groupby as _groupby
        sorted_mvs = sorted(movements, key=lambda m: m.get("category", ""))
        for cat, grp in _groupby(sorted_mvs, key=lambda m: m.get("category", "")):
            # Category sub-header
            ws_icb.merge_cells(f"A{icb_row}:G{icb_row}")
            ch = ws_icb.cell(row=icb_row, column=1, value=cat)
            ch.font = body_font(bold=True, color="1A2B4C")
            ch.fill = fill(C_SUBHEAD)
            ws_icb.row_dimensions[icb_row].height = 16
            icb_row += 1
            for m in grp:
                var = m.get("variance")
                is_fav_m = m.get("is_fav", True)
                row_bg = C_FAV_BG if (is_fav_m and var) else (C_ADV_BG if var else C_WHITE)
                vals = [
                    m.get("account", ""),
                    m.get("category", ""),
                    m.get("value"),
                    m.get("prior_value"),
                    var,
                    m.get("variance_pct"),
                    "",   # blank narrative for accountant to fill
                ]
                for ci, v in enumerate(vals, 1):
                    c2 = ws_icb.cell(row=icb_row, column=ci, value=gbp_fmt(v) if ci in (3,4,5) else (pct_fmt(v) if ci == 6 else v))
                    c2.font   = body_font()
                    c2.fill   = fill(row_bg)
                    c2.border = thin_border()
                    c2.alignment = Alignment(horizontal="right" if ci > 2 else "left", vertical="center", wrap_text=(ci == 7))
                    if ci in (3, 4) and isinstance(v, (int, float)):
                        c2.number_format = GBP_FMT
                    elif ci == 5 and isinstance(v, (int, float)):
                        c2.number_format = GBP_SFMT
                    elif ci == 6 and isinstance(v, (int, float)):
                        c2.number_format = PCT_FMT
                icb_row += 1

        ws_icb.column_dimensions["A"].width = 36
        ws_icb.column_dimensions["B"].width = 18
        ws_icb.column_dimensions["C"].width = 15
        ws_icb.column_dimensions["D"].width = 15
        ws_icb.column_dimensions["E"].width = 15
        ws_icb.column_dimensions["F"].width = 10
        ws_icb.column_dimensions["G"].width = 40

        # ── Workforce Sheet ───────────────────────────────────────────────────
        wb_data = nhs_data.get("workforce_breakdown", {})
        if wb_data:
            ws_wf = wb.create_sheet("Workforce")
            ws_wf.merge_cells("A1:D1")
            wf_t = ws_wf["A1"]
            wf_t.value = f"Workforce Cost Breakdown — {period_label_str}"
            wf_t.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
            wf_t.alignment = Alignment(horizontal="left", vertical="center")
            ws_wf.row_dimensions[1].height = 22
            for ci, h in enumerate(["Role Type", "Total (£)", "% of Payroll", "Accounts"], 1):
                c = ws_wf.cell(row=2, column=ci, value=h)
                c.font  = hdr_font()
                c.fill  = fill(C_HEADER)
                c.alignment = Alignment(horizontal="left", vertical="center")
            total_payroll = sum(abs(v.get("total", 0)) for v in wb_data.values())
            wf_row = 3
            for role_key, role_label in [("clinical","Clinical"),("locum","Locum (Sessional)"),("management","Management"),("admin","Admin"),("other","Other")]:
                rd = wb_data.get(role_key, {})
                amt = abs(rd.get("total", 0))
                if not amt:
                    continue
                pct_val = round(amt / total_payroll * 100, 1) if total_payroll else 0
                accs = ", ".join(rd.get("accounts", [])[:5])
                for ci, v in enumerate([role_label, amt, pct_val, accs], 1):
                    c = ws_wf.cell(row=wf_row, column=ci, value=v)
                    c.font   = body_font()
                    c.fill   = fill(C_SUBHEAD if wf_row % 2 == 0 else C_WHITE)
                    c.border = thin_border()
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    if ci == 2 and isinstance(v, (int, float)):
                        c.number_format = GBP_FMT
                    elif ci == 3 and isinstance(v, (int, float)):
                        c.number_format = "0.0%"
                        c.value = v / 100
                wf_row += 1
            ws_wf.column_dimensions["A"].width = 22
            ws_wf.column_dimensions["B"].width = 15
            ws_wf.column_dimensions["C"].width = 15
            ws_wf.column_dimensions["D"].width = 50

        # ── Partner Accounts Sheet ────────────────────────────────────────────
        partner_drawings = nhs_data.get("partner_drawings", 0) or 0
        wte_partners = nhs_data.get("wte_partners", 0) or 0
        net_profit = next((k["value"] for k in kpis if k.get("icon") == "wallet"), None)
        if net_profit is not None:
            ws_pa = wb.create_sheet("Partner Accounts")
            ws_pa.merge_cells("A1:C1")
            pa_t = ws_pa["A1"]
            pa_t.value = f"Partner Accounts Summary — {period_label_str}"
            pa_t.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
            pa_t.alignment = Alignment(horizontal="left", vertical="center")
            ws_pa.row_dimensions[1].height = 22
            available = net_profit - partner_drawings
            tax_reserve = max(0, round(available * 0.20, 0))
            pa_rows = [
                ("Net profit for period",        net_profit,        False),
                ("Less: partner drawings",        -partner_drawings, False),
                ("Surplus available to partners", available,         True),
                ("Suggested tax reserve (20%)",   -tax_reserve,      False),
            ]
            if wte_partners:
                pa_rows.append(("WTE partners", wte_partners, False))
                if wte_partners > 0:
                    pa_rows.append(("Profit per partner", round(net_profit / wte_partners, 2), False))
            for ri, (label, val, is_total) in enumerate(pa_rows, start=2):
                label_c = ws_pa.cell(row=ri, column=1, value=label)
                label_c.font = body_font(bold=is_total)
                label_c.fill = fill(C_SUBHEAD if is_total else C_WHITE)
                val_c = ws_pa.cell(row=ri, column=2, value=val)
                val_c.font = body_font(bold=is_total, color=C_FAV_FG if (isinstance(val,(int,float)) and val > 0) else C_ADV_FG)
                val_c.fill = fill(C_SUBHEAD if is_total else C_WHITE)
                val_c.number_format = GBP_FMT if isinstance(val, (int, float)) and abs(val) > 1 else "General"
                for cc in (label_c, val_c):
                    cc.border = thin_border()
                    cc.alignment = Alignment(horizontal="left" if cc.column == 1 else "right", vertical="center")
            ws_pa.column_dimensions["A"].width = 36
            ws_pa.column_dimensions["B"].width = 18

        # ── AISMA Mapping Sheet ───────────────────────────────────────────────
        ws_am = wb.create_sheet("AISMA Mapping")
        ws_am.merge_cells("A1:C1")
        am_t = ws_am["A1"]
        am_t.value = "AISMA Standard Chart of Accounts Mapping"
        am_t.font  = Font(name="Calibri", bold=True, size=13, color=C_HEADER)
        am_t.alignment = Alignment(horizontal="left", vertical="center")
        ws_am.row_dimensions[1].height = 22
        for ci, h in enumerate(["AISMA Heading", "MonthEndIQ Category", "Notes"], 1):
            c = ws_am.cell(row=2, column=ci, value=h)
            c.font  = hdr_font()
            c.fill  = fill(C_HEADER)
            c.alignment = Alignment(horizontal="left", vertical="center")
        aisma_map = [
            ("GMS / PMS / APMS Contract Income",   "Revenue",           "Global sum, MPIG, contract payments"),
            ("QOF Income",                          "Revenue",           "Aspiration and achievement payments"),
            ("Enhanced Services",                   "Revenue",           "DES, LES, PCN DES payments"),
            ("Other NHS Income",                    "Revenue",           "ICB payments, reimbursements"),
            ("Private / Non-NHS Income",            "Revenue",           "Private fees, non-NHS work"),
            ("Partner Drawings / Salaries",         "Staff Costs",       "GP partner drawings or salaried GP costs"),
            ("Locum / Sessional Costs",             "Staff Costs",       "Locum fees, sessional GP payments"),
            ("ARRS Staff Costs",                    "Staff Costs",       "Additional Roles Reimbursement Scheme"),
            ("Other Clinical Staff",                "Staff Costs",       "Nurses, ANPs, HCAs, pharmacists"),
            ("Admin & Management Staff",            "Staff Costs / Management Costs", "Receptionists, practice manager"),
            ("Employer NI & Pension",               "Staff Costs",       "Including NHS superannuation"),
            ("Drugs & Medical Supplies",            "Direct Costs",      "Dispensing costs, clinical consumables"),
            ("Premises (Rent, Rates, Utilities)",   "Premises Costs",    "Including notional rent reimbursements"),
            ("Clinical Systems & IT",               "IT Costs",          "EMIS, SystmOne, Docman, etc."),
            ("Medical Indemnity",                   "Insurance",         "MPS, MDDUS, MDU subscriptions"),
            ("CQC Registration Fee",                "Professional Fees", "Annual CQC fee"),
            ("Accountancy & Legal",                 "Professional Fees", "External professional services"),
            ("Finance Costs",                       "Finance Costs",     "Bank charges, loan interest"),
            ("Depreciation",                        "Depreciation & Amortisation", "Equipment and fixtures"),
        ]
        for ri, (aisma_h, meiq_cat, notes) in enumerate(aisma_map, start=3):
            for ci, v in enumerate([aisma_h, meiq_cat, notes], 1):
                c = ws_am.cell(row=ri, column=ci, value=v)
                c.font   = body_font()
                c.fill   = fill(C_SUBHEAD if ri % 2 == 0 else C_WHITE)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws_am.column_dimensions["A"].width = 38
        ws_am.column_dimensions["B"].width = 28
        ws_am.column_dimensions["C"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
